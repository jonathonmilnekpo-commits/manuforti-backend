import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=14, color='FFFFFF')
header_fill = PatternFill(start_color='002147', end_color='002147', fill_type='solid')
subheader_font = Font(bold=True, size=12)
label_font = Font(bold=True)
low_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
med_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Amber
high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== WORKSHEET 1: SUMMARY ====================
ws_summary = wb.active
ws_summary.title = 'Summary'

# Header
ws_summary['A1'] = 'PRODUCT 1 RISK SCORING SUMMARY'
ws_summary['A1'].font = Font(bold=True, size=16)
ws_summary.merge_cells('A1:F1')

# Supplier info
ws_summary['A3'] = 'Supplier Name:'
ws_summary['A3'].font = label_font
ws_summary['B3'] = '[Enter Supplier Name]'

ws_summary['A4'] = 'Analysis Date:'
ws_summary['A4'].font = label_font
ws_summary['B4'] = '=TODAY()'
ws_summary['B4'].number_format = 'YYYY-MM-DD'

ws_summary['A5'] = 'Analyst:'
ws_summary['A5'].font = label_font
ws_summary['B5'] = '[Agent Name]'

ws_summary['A6'] = 'Validation Status:'
ws_summary['A6'].font = label_font
ws_summary['B6'] = 'PENDING'

# Overall Score section
ws_summary['A8'] = 'OVERALL RISK SCORE'
ws_summary['A8'].font = Font(bold=True, size=14)
ws_summary['B8'] = '=ROUND(SUMPRODUCT(B14:B18,C14:C18)/100,0)'
ws_summary['B8'].font = Font(bold=True, size=24)
ws_summary['B8'].alignment = Alignment(horizontal='center')

ws_summary['A9'] = 'OVERALL RATING'
ws_summary['A9'].font = label_font
ws_summary['B9'] = '=IF(B8<=33,"LOW",IF(B8<=66,"MEDIUM","HIGH"))'
ws_summary['B9'].font = Font(bold=True, size=14)

ws_summary['A10'] = 'RECOMMENDATION'
ws_summary['A10'].font = label_font
ws_summary['B10'] = '=IF(B8<=33,"APPROVE",IF(B8<=66,"APPROVE WITH CONDITIONS","REJECT"))'
ws_summary['B10'].font = Font(bold=True, size=12)

# Pillar breakdown header
ws_summary['A12'] = 'PILLAR BREAKDOWN'
ws_summary['A12'].font = Font(bold=True, size=12, color='FFFFFF')
ws_summary['A12'].fill = header_fill
ws_summary.merge_cells('A12:E12')

# Column headers
ws_summary['A13'] = 'Risk Pillar'
ws_summary['B13'] = 'Score (0-100)'
ws_summary['C13'] = 'Weight %'
ws_summary['D13'] = 'Weighted Score'
ws_summary['E13'] = 'Rating'
for col in ['A13', 'B13', 'C13', 'D13', 'E13']:
    ws_summary[col].font = Font(bold=True)
    ws_summary[col].border = thin_border

# Pillar rows - using formulas that reference other sheets
pillars = [
    ('Financial', 35),
    ('Operational', 25),
    ('Geopolitical', 20),
    ('ESG', 15),
    ('Commercial', 5)
]

for i, (pillar, weight) in enumerate(pillars, start=14):
    ws_summary.cell(row=i, column=1, value=pillar).border = thin_border
    # Score will be entered manually or linked from other sheets
    ws_summary.cell(row=i, column=2, value='[Score]').border = thin_border
    ws_summary.cell(row=i, column=3, value=weight).border = thin_border
    ws_summary.cell(row=i, column=4, value=f'=B{i}*C{i}/100').border = thin_border
    ws_summary.cell(row=i, column=5, value=f'=IF(B{i}<=33,"LOW",IF(B{i}<=66,"MEDIUM","HIGH"))').border = thin_border

# Conditions section
ws_summary['A20'] = 'APPROVAL CONDITIONS'
ws_summary['A20'].font = Font(bold=True, size=12, color='FFFFFF')
ws_summary['A20'].fill = header_fill
ws_summary.merge_cells('A20:E20')

ws_summary['A21'] = '1.'
ws_summary['B21'] = '[Condition 1 - auto-populated if MEDIUM/HIGH]'
ws_summary['A22'] = '2.'
ws_summary['B22'] = '[Condition 2]'
ws_summary['A23'] = '3.'
ws_summary['B23'] = '[Condition 3]'

# Adjust column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 20
ws_summary.column_dimensions['C'].width = 12
ws_summary.column_dimensions['D'].width = 15
ws_summary.column_dimensions['E'].width = 12

# ==================== WORKSHEET 2: FINANCIAL ====================
ws_fin = wb.create_sheet('Financial')

ws_fin['A1'] = 'FINANCIAL RISK SCORING (35% Weight)'
ws_fin['A1'].font = Font(bold=True, size=14)
ws_fin.merge_cells('A1:H1')

# Headers
headers = ['Metric', 'Raw Value', 'Score (0-100)', 'Weight %', 'Weighted Score', 'LOW Threshold', 'MEDIUM Threshold', 'HIGH Threshold']
for col, header in enumerate(headers, 1):
    cell = ws_fin.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

# Financial metrics data
financial_data = [
    ('Debt/EBITDA', '1.4x', 30, '< 2.0x', '2.0x - 4.0x', '> 4.0x'),
    ('Net Cash Position', '+$85M', 25, 'Positive', 'Neutral/Small debt', 'Significant net debt'),
    ('EBITDA Margin', '13.3%', 25, '> 12%', '5% - 12%', '< 5% or negative'),
    ('Revenue CAGR (3yr)', '8.5%', 15, '> 5%', '0% - 5%', 'Negative'),
    ('Interest Coverage', '4.2x', 5, '> 3.0x', '1.5x - 3.0x', '< 1.5x')
]

for i, (metric, raw_val, weight, low_thresh, med_thresh, high_thresh) in enumerate(financial_data, start=4):
    ws_fin.cell(row=i, column=1, value=metric).border = thin_border
    ws_fin.cell(row=i, column=2, value=raw_val).border = thin_border
    ws_fin.cell(row=i, column=3, value='[Enter Score 0-100]').border = thin_border
    ws_fin.cell(row=i, column=4, value=weight).border = thin_border
    ws_fin.cell(row=i, column=5, value=f'=C{i}*D{i}/100').border = thin_border
    ws_fin.cell(row=i, column=6, value=low_thresh).border = thin_border
    ws_fin.cell(row=i, column=7, value=med_thresh).border = thin_border
    ws_fin.cell(row=i, column=8, value=high_thresh).border = thin_border

# Total row
ws_fin['A10'] = 'FINANCIAL PILLAR SCORE'
ws_fin['A10'].font = Font(bold=True)
ws_fin['D10'] = '=SUM(D4:D8)'
ws_fin['D10'].font = Font(bold=True)
ws_fin['E10'] = '=SUM(E4:E8)'
ws_fin['E10'].font = Font(bold=True, size=14)
ws_fin['F10'] = '=IF(E10<=33,"LOW",IF(E10<=66,"MEDIUM","HIGH"))'
ws_fin['F10'].font = Font(bold=True)

# Score guidance
ws_fin['A13'] = 'SCORING GUIDANCE:'
ws_fin['A13'].font = Font(bold=True, italic=True)
ws_fin['A14'] = '0-33 (LOW): Strong financial position, low leverage, healthy margins'
ws_fin['A15'] = '34-66 (MEDIUM): Some financial pressure, manageable but requires monitoring'
ws_fin['A16'] = '67-100 (HIGH): Financial distress, high leverage, or declining performance'

# Adjust column widths
for i, width in enumerate([25, 15, 15, 12, 15, 18, 20, 20], 1):
    ws_fin.column_dimensions[get_column_letter(i)].width = width

# ==================== WORKSHEET 3: OPERATIONAL ====================
ws_ops = wb.create_sheet('Operational')

ws_ops['A1'] = 'OPERATIONAL RISK SCORING (25% Weight)'
ws_ops['A1'].font = Font(bold=True, size=14)
ws_ops.merge_cells('A1:H1')

# Headers
for col, header in enumerate(headers, 1):
    cell = ws_ops.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

# Operational metrics
ops_data = [
    ('Order Book Coverage', '>18 months', 25, '> 18 months', '12-18 months', '< 12 months'),
    ('Capacity Utilization', '75%', 20, '70-85%', '85-95% or <70%', '>95% or <50%'),
    ('Quality Certifications', 'ISO 9001 + Industry', 20, 'ISO 9001 + industry cert', 'ISO 9001 only', 'No certifications'),
    ('Supply Chain Concentration', 'Diversified', 20, '>3 regions', '2-3 regions', '1 region'),
    ('Delivery Track Record', '97% on-time', 15, '>95% on-time', '85-95% on-time', '<85% on-time')
]

for i, (metric, raw_val, weight, low_thresh, med_thresh, high_thresh) in enumerate(ops_data, start=4):
    ws_ops.cell(row=i, column=1, value=metric).border = thin_border
    ws_ops.cell(row=i, column=2, value=raw_val).border = thin_border
    ws_ops.cell(row=i, column=3, value='[Enter Score 0-100]').border = thin_border
    ws_ops.cell(row=i, column=4, value=weight).border = thin_border
    ws_ops.cell(row=i, column=5, value=f'=C{i}*D{i}/100').border = thin_border
    ws_ops.cell(row=i, column=6, value=low_thresh).border = thin_border
    ws_ops.cell(row=i, column=7, value=med_thresh).border = thin_border
    ws_ops.cell(row=i, column=8, value=high_thresh).border = thin_border

ws_ops['A10'] = 'OPERATIONAL PILLAR SCORE'
ws_ops['A10'].font = Font(bold=True)
ws_ops['D10'] = '=SUM(D4:D8)'
ws_ops['D10'].font = Font(bold=True)
ws_ops['E10'] = '=SUM(E4:E8)'
ws_ops['E10'].font = Font(bold=True, size=14)
ws_ops['F10'] = '=IF(E10<=33,"LOW",IF(E10<=66,"MEDIUM","HIGH"))'
ws_ops['F10'].font = Font(bold=True)

for i, width in enumerate([30, 20, 15, 12, 15, 25, 25, 20], 1):
    ws_ops.column_dimensions[get_column_letter(i)].width = width

# ==================== WORKSHEET 4: GEOPOLITICAL ====================
ws_geo = wb.create_sheet('Geopolitical')

ws_geo['A1'] = 'GEOPOLITICAL RISK SCORING (20% Weight)'
ws_geo['A1'].font = Font(bold=True, size=14)
ws_geo.merge_cells('A1:H1')

for col, header in enumerate(headers, 1):
    cell = ws_geo.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

geo_data = [
    ('Country Risk Rating', 'Low Risk', 35, 'BMI/Fitch: Low', 'BMI/Fitch: Medium', 'BMI/Fitch: High'),
    ('Regulatory Stability', 'Stable', 25, 'Stable, predictable', 'Some uncertainty', 'High volatility'),
    ('Trade/Sanctions Exposure', 'None', 25, 'No exposure', 'Minor indirect', 'Direct sanctions risk'),
    ('Currency Volatility', '5% annual', 15, '<10% annual', '10-25% annual', '>25% annual')
]

for i, (metric, raw_val, weight, low_thresh, med_thresh, high_thresh) in enumerate(geo_data, start=4):
    ws_geo.cell(row=i, column=1, value=metric).border = thin_border
    ws_geo.cell(row=i, column=2, value=raw_val).border = thin_border
    ws_geo.cell(row=i, column=3, value='[Enter Score 0-100]').border = thin_border
    ws_geo.cell(row=i, column=4, value=weight).border = thin_border
    ws_geo.cell(row=i, column=5, value=f'=C{i}*D{i}/100').border = thin_border
    ws_geo.cell(row=i, column=6, value=low_thresh).border = thin_border
    ws_geo.cell(row=i, column=7, value=med_thresh).border = thin_border
    ws_geo.cell(row=i, column=8, value=high_thresh).border = thin_border

ws_geo['A9'] = 'GEOPOLITICAL PILLAR SCORE'
ws_geo['A9'].font = Font(bold=True)
ws_geo['D9'] = '=SUM(D4:D7)'
ws_geo['D9'].font = Font(bold=True)
ws_geo['E9'] = '=SUM(E4:E7)'
ws_geo['E9'].font = Font(bold=True, size=14)
ws_geo['F9'] = '=IF(E9<=33,"LOW",IF(E9<=66,"MEDIUM","HIGH"))'
ws_geo['F9'].font = Font(bold=True)

for i, width in enumerate([30, 20, 15, 12, 15, 25, 25, 20], 1):
    ws_geo.column_dimensions[get_column_letter(i)].width = width

# ==================== WORKSHEET 5: ESG ====================
ws_esg = wb.create_sheet('ESG')

ws_esg['A1'] = 'ESG RISK SCORING (15% Weight)'
ws_esg['A1'].font = Font(bold=True, size=14)
ws_esg.merge_cells('A1:H1')

for col, header in enumerate(headers, 1):
    cell = ws_esg.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

esg_data = [
    ('Environmental (E)', 'EcoVadis B', 35, 'EcoVadis A/B, ISO 14001', 'EcoVadis C, no major violations', 'EcoVadis D/E, violations'),
    ('Social (S)', 'Strong CoC', 35, 'Strong CoC, no controversies', 'Minor controversies resolved', 'Active litigation, labor issues'),
    ('Governance (G)', 'Transparent', 30, 'Transparent ownership', 'Complex structure', 'Sanctions, corruption cases')
]

for i, (metric, raw_val, weight, low_thresh, med_thresh, high_thresh) in enumerate(esg_data, start=4):
    ws_esg.cell(row=i, column=1, value=metric).border = thin_border
    ws_esg.cell(row=i, column=2, value=raw_val).border = thin_border
    ws_esg.cell(row=i, column=3, value='[Enter Score 0-100]').border = thin_border
    ws_esg.cell(row=i, column=4, value=weight).border = thin_border
    ws_esg.cell(row=i, column=5, value=f'=C{i}*D{i}/100').border = thin_border
    ws_esg.cell(row=i, column=6, value=low_thresh).border = thin_border
    ws_esg.cell(row=i, column=7, value=med_thresh).border = thin_border
    ws_esg.cell(row=i, column=8, value=high_thresh).border = thin_border

ws_esg['A8'] = 'ESG PILLAR SCORE'
ws_esg['A8'].font = Font(bold=True)
ws_esg['D8'] = '=SUM(D4:D6)'
ws_esg['D8'].font = Font(bold=True)
ws_esg['E8'] = '=SUM(E4:E6)'
ws_esg['E8'].font = Font(bold=True, size=14)
ws_esg['F8'] = '=IF(E8<=33,"LOW",IF(E8<=66,"MEDIUM","HIGH"))'
ws_esg['F8'].font = Font(bold=True)

ws_esg['A11'] = 'ESG OVERRIDE RULE:'
ws_esg['A11'].font = Font(bold=True, color='FF0000')
ws_esg['A12'] = 'If ANY ESG sub-pillar is HIGH (controversy/litigation), overall ESG = HIGH regardless of other scores.'

for i, width in enumerate([25, 20, 15, 12, 15, 30, 30, 25], 1):
    ws_esg.column_dimensions[get_column_letter(i)].width = width

# ==================== WORKSHEET 6: COMMERCIAL ====================
ws_comm = wb.create_sheet('Commercial')

ws_comm['A1'] = 'COMMERCIAL RISK SCORING (5% Weight)'
ws_comm['A1'].font = Font(bold=True, size=14)
ws_comm.merge_cells('A1:H1')

for col, header in enumerate(headers, 1):
    cell = ws_comm.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

comm_data = [
    ('Pricing Volatility', 'Stable', 35, 'Stable, predictable', 'Moderate fluctuations', 'Highly volatile'),
    ('Contract Flexibility', 'Standard', 35, 'Standard terms OK', 'Some negotiation needed', 'Aggressive terms'),
    ('Market Position', 'Leader', 30, 'Leader / strong #2', 'Mid-tier', 'Weak, struggling')
]

for i, (metric, raw_val, weight, low_thresh, med_thresh, high_thresh) in enumerate(comm_data, start=4):
    ws_comm.cell(row=i, column=1, value=metric).border = thin_border
    ws_comm.cell(row=i, column=2, value=raw_val).border = thin_border
    ws_comm.cell(row=i, column=3, value='[Enter Score 0-100]').border = thin_border
    ws_comm.cell(row=i, column=4, value=weight).border = thin_border
    ws_comm.cell(row=i, column=5, value=f'=C{i}*D{i}/100').border = thin_border
    ws_comm.cell(row=i, column=6, value=low_thresh).border = thin_border
    ws_comm.cell(row=i, column=7, value=med_thresh).border = thin_border
    ws_comm.cell(row=i, column=8, value=high_thresh).border = thin_border

ws_comm['A8'] = 'COMMERCIAL PILLAR SCORE'
ws_comm['A8'].font = Font(bold=True)
ws_comm['D8'] = '=SUM(D4:D6)'
ws_comm['D8'].font = Font(bold=True)
ws_comm['E8'] = '=SUM(E4:E6)'
ws_comm['E8'].font = Font(bold=True, size=14)
ws_comm['F8'] = '=IF(E8<=33,"LOW",IF(E8<=66,"MEDIUM","HIGH"))'
ws_comm['F8'].font = Font(bold=True)

for i, width in enumerate([25, 20, 15, 12, 15, 25, 25, 20], 1):
    ws_comm.column_dimensions[get_column_letter(i)].width = width

# ==================== WORKSHEET 7: RAW DATA ====================
ws_raw = wb.create_sheet('Raw Data')

ws_raw['A1'] = 'RAW DATA INPUTS'
ws_raw['A1'].font = Font(bold=True, size=14)
ws_raw.merge_cells('A1:D1')

ws_raw['A3'] = 'Field'
ws_raw['B3'] = 'Value'
ws_raw['C3'] = 'Source'
ws_raw['D3'] = 'Date Sourced'
for col in ['A3', 'B3', 'C3', 'D3']:
    ws_raw[col].font = Font(bold=True)
    ws_raw[col].fill = header_fill
    ws_raw[col].font = Font(bold=True, color='FFFFFF')

# Raw data fields
raw_fields = [
    'revenue_2024', 'revenue_2023', 'revenue_2022',
    'revenue_yoy', 'ebitda', 'ebitda_margin',
    'net_profit', 'cagr_3yr', 'order_book',
    'gross_debt', 'net_cash', 'debt_ebitda',
    'headquarters', 'employees', 'founded_year',
    'ceo_name', 'ceo_tenure', 'parent_company',
    'esg_environmental', 'esg_social', 'esg_governance',
    'ecovadis_rating', 'iso_14001', 'iso_9001',
    'controversy_1', 'controversy_2', 'controversy_3',
    'competitor_1', 'competitor_1_revenue',
    'competitor_2', 'competitor_2_revenue',
    'competitor_3', 'competitor_3_revenue'
]

for i, field in enumerate(raw_fields, start=4):
    ws_raw.cell(row=i, column=1, value=field)
    ws_raw.cell(row=i, column=2, value='[Enter value]')
    ws_raw.cell(row=i, column=3, value='[Source]')
    ws_raw.cell(row=i, column=4, value='[Date]')

ws_raw.column_dimensions['A'].width = 25
ws_raw.column_dimensions['B'].width = 20
ws_raw.column_dimensions['C'].width = 25
ws_raw.column_dimensions['D'].width = 15

# Save workbook
output_path = '/Users/jonathonmilne/.openclaw/workspace/templates/product1_scoring_template.xlsx'
wb.save(output_path)
print(f'Excel template created: {output_path}')
