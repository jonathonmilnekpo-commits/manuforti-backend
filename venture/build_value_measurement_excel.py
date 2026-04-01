import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side, numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import datetime

wb = openpyxl.Workbook()

# ── COLOUR PALETTE ─────────────────────────────────────────────────────────────
NAVY     = "002147"
STEEL    = "2B6CB0"
LTBLUE   = "EBF4FF"
MID_GREY = "718096"
WHITE    = "FFFFFF"
AMBER    = "D97F06"
GREEN    = "27AE60"
RED      = "C0392B"
DARK     = "2D3A4A"
PALE_GRN = "D4EDDA"
PALE_RED = "F8D7DA"
PALE_AMB = "FFF3CD"

def nfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def nfont(bold=False, color=DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def nalign(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def nborder(sides="all", style="thin"):
    s = Side(style=style)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    elif sides == "bottom":
        return Border(bottom=s)
    elif sides == "outer":
        return Border(left=s, right=s, top=s, bottom=s)
    return Border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_row(ws, row, cols_data, bg=NAVY, fg=WHITE, size=10, bold=True):
    for col, (val, width) in enumerate(cols_data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = nfill(bg)
        c.font = nfont(bold=bold, color=fg, size=size)
        c.alignment = nalign("center", "center", wrap=True)
        c.border = nborder()
        if width:
            set_col_width(ws, col, width)

def data_cell(ws, row, col, value, bg=WHITE, fg=DARK, bold=False, number_format=None,
              align="left", wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = nfill(bg)
    c.font = nfont(bold=bold, color=fg, size=size)
    c.alignment = nalign(align, "center", wrap=wrap)
    c.border = nborder()
    if number_format:
        c.number_format = number_format
    return c

def section_title(ws, row, title, cols=10, bg=STEEL, fg=WHITE):
    c = ws.cell(row=row, column=1, value=title)
    c.fill = nfill(bg)
    c.font = nfont(bold=True, color=fg, size=11)
    c.alignment = nalign("left", "center")
    c.border = nborder()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Dashboard"
ws1.sheet_view.showGridLines = False

# Title banner
ws1.merge_cells("A1:L2")
c = ws1["A1"]
c.value = "STATKRAFT PROCUREMENT AI — VALUE MEASUREMENT DASHBOARD"
c.fill = nfill(NAVY)
c.font = Font(bold=True, color=WHITE, size=16, name="Calibri")
c.alignment = nalign("center", "center")

ws1.merge_cells("A3:L3")
c = ws1["A3"]
c.value = f"Last Updated: {datetime.date.today().strftime('%d %B %Y')}  |  All values in EUR unless stated  |  Statkraft Procurement AI Programme"
c.fill = nfill(STEEL)
c.font = nfont(color=WHITE, size=10, italic=True)
c.alignment = nalign("center", "center")

ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 18

# KPI Summary row
ws1["A5"] = "KPI SUMMARY"
ws1["A5"].fill = nfill(STEEL)
ws1["A5"].font = nfont(bold=True, color=WHITE, size=11)
ws1.merge_cells("A5:L5")
ws1["A5"].alignment = nalign("left")

kpi_headers = [
    ("Initiative",18),("Category",14),("Status",10),("Baseline KPI",14),
    ("Target KPI",14),("Current KPI",14),("% Achieved",10),("Est. EUR Saved",14),
    ("ROI",8),("Go-Live Date",12),("Next Review",12),("RAG",6)
]
header_row(ws1, 6, kpi_headers)

# Sample data rows (formulas)
initiatives = [
    ["Spend Classification",    "Data Quality",    "🟡 Pilot",    "65% accuracy",  "85% accuracy",  "",  "",  "=IF(H7<>\"\",H7,\"TBD\")",  "",  "2026-04-01",  "2026-07-01",  "🟡"],
    ["Contract NLP Scanning",   "Contract Mgmt",   "🟡 Pilot",    "Manual review", "80% faster",    "",  "",  "=IF(H8<>\"\",H8,\"TBD\")",  "",  "2026-04-15",  "2026-07-15",  "🟡"],
    ["Supplier Monitoring",     "Risk Mgmt",       "🟡 Pilot",    "Annual review", "Continuous",    "",  "",  "=IF(H9<>\"\",H9,\"TBD\")",  "",  "2026-05-01",  "2026-08-01",  "🟡"],
    ["RFQ Drafter",             "Sourcing",        "🔵 Planned",  "8 hrs/event",   "2 hrs/event",   "",  "",  "",  "",  "2026-07-01",  "2026-10-01",  "🔵"],
    ["Demand Forecasting MRO",  "MRO",             "🔵 Planned",  "Reactive only", "6-12m forecast","",  "",  "",  "",  "2026-07-15",  "2026-10-15",  "🔵"],
    ["Negotiation Pack Builder","Sourcing",        "🔵 Planned",  "Manual prep",   "AI-assisted",   "",  "",  "",  "",  "2026-08-01",  "2026-11-01",  "🔵"],
    ["Category Strategist",     "Strategy",        "🔵 Planned",  "Weeks to draft","Days to draft", "",  "",  "",  "",  "2026-09-01",  "2026-12-01",  "🔵"],
    ["CSRD Compliance Agent",   "Compliance",      "🔵 Planned",  "Manual checks", "Automated",     "",  "",  "",  "",  "2026-10-01",  "2027-01-01",  "🔵"],
]

row_colors = [LTBLUE, WHITE]
for i, row_data in enumerate(initiatives):
    row = 7 + i
    bg = row_colors[i % 2]
    for col, val in enumerate(row_data, 1):
        if col in [8, 9] and "=IF" in str(val):
            c = ws1.cell(row=row, column=col, value=val)
            c.fill = nfill(bg)
            c.font = nfont()
            c.border = nborder()
            c.alignment = nalign("center")
        else:
            data_cell(ws1, row, col, val, bg=bg, align="center" if col > 2 else "left")

# Totals row
tot_row = 7 + len(initiatives)
ws1.merge_cells(f"A{tot_row}:G{tot_row}")
c = ws1.cell(row=tot_row, column=1, value="PROGRAMME TOTAL")
c.fill = nfill(NAVY); c.font = nfont(bold=True, color=WHITE); c.alignment = nalign("right"); c.border = nborder()

for col in range(1, 13):
    c = ws1.cell(row=tot_row, column=col)
    c.fill = nfill(NAVY); c.border = nborder()
    if col == 1:
        c.value = "PROGRAMME TOTAL"; c.font = nfont(bold=True, color=WHITE, size=11)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: VALUE CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("💰 Value Calculator")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:J2")
c = ws2["A1"]
c.value = "PROCUREMENT AI VALUE CALCULATOR  —  Enter yellow cells. Green cells calculate automatically."
c.fill = nfill(NAVY); c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
c.alignment = nalign("center", "center")
ws2.row_dimensions[1].height = 28

# ── SECTION: SPEND & TEAM INPUTS ──────────────────────────────────────────────
section_title(ws2, 4, "  SECTION 1: YOUR INPUTS (change yellow cells)", cols=10)

inputs = [
    ("Total annual procurement spend (EUR)", 500_000_000, "EUR", "B5"),
    ("Number of sourcing events per year", 250, "events/yr", "B6"),
    ("Average hours per sourcing event (RFQ to award)", 40, "hrs", "B7"),
    ("Number of contracts in portfolio", 1200, "contracts", "B8"),
    ("Hours to manually review one contract", 4, "hrs", "B9"),
    ("Number of suppliers actively monitored", 200, "suppliers", "B10"),
    ("Annual cost of manual supplier reviews", 150_000, "EUR/yr", "B11"),
    ("Procurement headcount (FTEs)", 45, "FTEs", "B12"),
    ("Average fully-loaded FTE cost (EUR/yr)", 120_000, "EUR/yr", "B13"),
    ("Current spend classification accuracy", 0.65, "%", "B14"),
]

header_row(ws2, 3, [("Input Parameter", 40), ("Your Value", 18), ("Unit", 15), ("Source / Assumption", 30)], bg=STEEL)

for i, (label, val, unit, _) in enumerate(inputs):
    row = 4 + i
    data_cell(ws2, row, 1, label, bg=WHITE)
    c = ws2.cell(row=row, column=2, value=val)
    c.fill = nfill("FFFACD"); c.font = nfont(bold=True, color=DARK, size=11)
    c.alignment = nalign("center"); c.border = nborder()
    if "%" in unit and val < 2:
        c.number_format = "0%"
    elif val > 1000:
        c.number_format = '#,##0'
    data_cell(ws2, row, 3, unit, bg=LTBLUE, align="center")
    # Source notes
    sources = ["Statkraft internal","Statkraft estimate","Industry avg","Statkraft estimate",
               "Industry avg (Luminance benchmark)","Statkraft target","Statkraft estimate",
               "Statkraft internal","Statkraft HR data","Current state"]
    data_cell(ws2, row, 4, sources[i], bg=WHITE, fg=MID_GREY)

ws2.row_dimensions[3].height = 22

# ── SECTION: VALUE OUTPUTS ────────────────────────────────────────────────────
section_title(ws2, 16, "  SECTION 2: CALCULATED VALUE  (auto-calculated from inputs above)", cols=10, bg=NAVY)

header_row(ws2, 17, [("Use Case",30),("Metric",20),("Current State",16),("With AI",16),
                      ("Annual Saving (EUR)",18),("ROI Assumption",18),(None,0),(None,0),(None,0),(None,0)], bg=STEEL)

# Rows 18 onwards = calculated outputs referencing Section 1
calc_rows = [
    ("RFQ Generation","Hours saved per event","=B7","=B7*0.75",
     "=(B7-B7*0.25)*B12*B6/B12*B13/1750","75% time reduction (Ørsted benchmark)"),
    ("Contract NLP Review","Hours per contract review","=B9","=B9*0.2",
     "=(B9-B9*0.2)*B8*B13/1750","80% faster (Luminance / Icertis benchmark)"),
    ("Supplier Monitoring","Annual monitoring cost","=B11","=B11*0.3",
     "=B11-B11*0.3","70% cost reduction (continuous vs annual review)"),
    ("Spend Classification","Classification accuracy","=B14","=0.88",
     "=B5*(0.88-B14)*0.05","5% spend recovery per accuracy point gained"),
    ("Demand Forecasting (MRO)","Emergency premium spend","=B5*0.03","=B5*0.03*0.5",
     "=B5*0.03*0.5","50% reduction in emergency buys (Vattenfall benchmark)"),
    ("Negotiation Preparation","Sourcing events with full data pack","=B6*0.2","=B6*0.8",
     "=B5*0.002","0.2% additional savings from better-prepared negotiations"),
    ("Category Strategy","Strategies completed per year","=3","=12",
     "=B5*0.005","0.5% strategic savings from better category coverage"),
    ("CSRD Compliance","Manual compliance checks","=B10*12","=B10*0.1",
     "=(B10*12-B10*0.1)*B13/1750*4","4 hrs per check vs 0.1 automated equiv."),
]

row_colors2 = [LTBLUE, WHITE]
for i, (name, metric, curr, ai_val, saving_formula, note) in enumerate(calc_rows):
    row = 18 + i
    bg = row_colors2[i % 2]
    data_cell(ws2, row, 1, name, bg=bg, bold=True)
    data_cell(ws2, row, 2, metric, bg=bg)
    # Current state (formula)
    c = ws2.cell(row=row, column=3, value=curr)
    c.fill = nfill(bg); c.font = nfont(); c.border = nborder(); c.alignment = nalign("center")
    try: c.number_format = '#,##0'
    except: pass
    # With AI (formula)
    c2 = ws2.cell(row=row, column=4, value=ai_val)
    c2.fill = nfill(PALE_GRN); c2.font = nfont(color=GREEN, bold=True)
    c2.border = nborder(); c2.alignment = nalign("center")
    try: c2.number_format = '#,##0'
    except: pass
    # Saving formula
    c3 = ws2.cell(row=row, column=5, value=saving_formula)
    c3.fill = nfill(PALE_GRN); c3.font = nfont(bold=True, color=GREEN, size=11)
    c3.border = nborder(); c3.alignment = nalign("center")
    c3.number_format = '€#,##0'
    data_cell(ws2, row, 6, note, bg=bg, fg=MID_GREY)

# Total saving
tot = 26
ws2.merge_cells(f"A{tot}:D{tot}")
c = ws2.cell(row=tot, column=1, value="TOTAL ESTIMATED ANNUAL VALUE")
c.fill = nfill(NAVY); c.font = nfont(bold=True, color=WHITE, size=12); c.border = nborder(); c.alignment = nalign("right")
c2 = ws2.cell(row=tot, column=5, value=f"=SUM(E18:E25)")
c2.fill = nfill(NAVY); c2.font = nfont(bold=True, color=WHITE, size=14)
c2.border = nborder(); c2.alignment = nalign("center"); c2.number_format = '€#,##0'

# Programme cost input
c28 = ws2.cell(row=28, column=1, value="Estimated programme cost (annual, EUR)")
c28.fill = nfill(WHITE)
c28.font = nfont(bold=True)
c28.border = nborder()
c = ws2.cell(row=28, column=5, value=750000)
c.fill = nfill("FFFACD"); c.font = nfont(bold=True, size=11); c.border = nborder()
c.alignment = nalign("center"); c.number_format = '€#,##0'

ws2.merge_cells("A29:D29")
c29 = ws2.cell(row=29, column=1, value="NET ANNUAL VALUE")
c29.fill = nfill(STEEL); c29.font = nfont(bold=True, color=WHITE, size=12)
c29.border = nborder(); c29.alignment = nalign("right")
c = ws2.cell(row=29, column=5, value="=E26-E28")
c.fill = nfill(STEEL); c.font = nfont(bold=True, color=WHITE, size=14)
c.border = nborder(); c.alignment = nalign("center"); c.number_format = '€#,##0'

ws2.merge_cells("A30:D30")
c30 = ws2.cell(row=30, column=1, value="ROI (x)")
c30.fill = nfill(WHITE); c30.font = nfont(bold=True); c30.border = nborder(); c30.alignment = nalign("right")
c = ws2.cell(row=30, column=5, value="=E26/E28")
c.fill = nfill(PALE_GRN); c.font = nfont(bold=True, color=GREEN, size=14)
c.border = nborder(); c.alignment = nalign("center"); c.number_format = '0.0"x"'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: USE CASE PRIORITISATION
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🎯 Use Case Prioritiser")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:M2")
c = ws3["A1"]
c.value = "USE CASE PRIORITISATION MATRIX  —  Score each use case 1–5 on each dimension. Priority Score auto-calculates."
c.fill = nfill(NAVY); c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = nalign("center", "center"); ws3.row_dimensions[1].height = 28

# Weights row
section_title(ws3, 4, "  SCORING WEIGHTS (adjust to reflect Statkraft priorities — must sum to 100%)", cols=13)
weights_labels = ["Business Value (40%)", "Implementation Speed (20%)", "Data Availability (15%)",
                  "Risk (15%)", "Strategic Fit (10%)"]
weights_vals   = [0.40, 0.20, 0.15, 0.15, 0.10]

for i, (lbl, val) in enumerate(zip(weights_labels, weights_vals)):
    col = 3 + i
    c = ws3.cell(row=5, column=col, value=val)
    c.fill = nfill("FFFACD"); c.font = nfont(bold=True, size=10)
    c.alignment = nalign("center"); c.border = nborder(); c.number_format = "0%"
    c2 = ws3.cell(row=6, column=col, value=lbl)
    c2.fill = nfill(LTBLUE); c2.font = nfont(size=9, color=DARK)
    c2.alignment = nalign("center", wrap=True); c2.border = nborder()

header_row(ws3, 7, [
    ("Use Case",30),("Category",14),
    ("Business\nValue (1-5)",14),("Implementation\nSpeed (1-5)",14),
    ("Data\nAvailability (1-5)",14),("Risk Score\n(1-5, 5=low risk)",14),
    ("Strategic\nFit (1-5)",14),("PRIORITY\nSCORE",12),("Rank",7),
    ("Recommended\nTrack",14),("Effort\n(weeks)",10),("Est. Saving\n(EUR/yr)",14),("Notes",22)
], bg=STEEL)

uc_data = [
    ("Spend Classification",       "Data Quality",   5,4,4,5,4,"=C8*$C$5+D8*$D$5+E8*$E$5+F8*$F$5+G8*$G$5", "=RANK(H8,H$8:H$19,0)", "Track B","4–6","=250000","Foundation for all other AI use cases"),
    ("Contract NLP Scanning",      "Contract Mgmt",  5,4,3,5,4,"=C9*$C$5+D9*$D$5+E9*$E$5+F9*$F$5+G9*$G$5",  "=RANK(H9,H$8:H$19,0)",  "Track B","6–8","=180000","Immediate portfolio value, no supplier risk"),
    ("Supplier Monitoring",        "Risk Mgmt",      5,3,4,4,5,"=C10*$C$5+D10*$D$5+E10*$E$5+F10*$F$5+G10*$G$5","=RANK(H10,H$8:H$19,0)","Track D","8–10","=150000","Manu Forti product — continuous monitoring"),
    ("RFQ Generation",             "Sourcing",       5,4,4,4,5,"=C11*$C$5+D11*$D$5+E11*$E$5+F11*$F$5+G11*$G$5","=RANK(H11,H$8:H$19,0)","Track D","8–12","=320000","4–8 hrs saved per event, proven at Ørsted/BP"),
    ("Demand Forecasting MRO",     "MRO",            4,3,3,4,4,"=C12*$C$5+D12*$D$5+E12*$E$5+F12*$F$5+G12*$G$5","=RANK(H12,H$8:H$19,0)","Track D","12–16","=450000","Vattenfall 6-12m horizon benchmark"),
    ("Negotiation Pack Builder",   "Sourcing",       5,3,4,4,4,"=C13*$C$5+D13*$D$5+E13*$E$5+F13*$F$5+G13*$G$5","=RANK(H13,H$8:H$19,0)","Track D","10–14","=250000","Market data + BATNA + walk-away"),
    ("Category Strategy Agent",    "Strategy",       5,2,4,3,5,"=C14*$C$5+D14*$D$5+E14*$E$5+F14*$F$5+G14*$G$5","=RANK(H14,H$8:H$19,0)","Track D","16–20","=500000","Highest value — weeks → days to draft"),
    ("CSRD Compliance Agent",      "Compliance",     4,3,3,3,5,"=C15*$C$5+D15*$D$5+E15*$E$5+F15*$F$5+G15*$G$5","=RANK(H15,H$8:H$19,0)","Track D","12–16","=200000","Mandatory for CSRD 2026 reporting"),
    ("Invoice Processing",         "P2P",            3,5,4,5,3,"=C16*$C$5+D16*$D$5+E16*$E$5+F16*$F$5+G16*$G$5","=RANK(H16,H$8:H$19,0)","Track B","4–6","=120000","3-way match + anomaly detection"),
    ("Market Price Intelligence",  "Market Intel",   4,5,5,5,3,"=C17*$C$5+D17*$D$5+E17*$E$5+F17*$F$5+G17*$G$5","=RANK(H17,H$8:H$19,0)","Track A","2–4","=80000","Daily briefings on commodity indices"),
    ("Supplier Onboarding",        "Supplier Mgmt",  3,3,3,4,3,"=C18*$C$5+D18*$D$5+E18*$E$5+F18*$F$5+G18*$G$5","=RANK(H18,H$8:H$19,0)","Track C","10–14","=100000","Automated due diligence"),
    ("Contract Drafting Assist",   "Contract Mgmt",  4,2,4,3,4,"=C19*$C$5+D19*$D$5+E19*$E$5+F19*$F$5+G19*$G$5","=RANK(H19,H$8:H$19,0)","Track C","14–18","=220000","40-60% legal cycle time reduction"),
]

row_colors3 = [LTBLUE, WHITE]
for i, row_data in enumerate(uc_data):
    row = 8 + i
    bg = row_colors3[i % 2]
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.border = nborder(); c.alignment = nalign("center" if col > 2 else "left", "center", wrap=(col==13))
        if col in [8]:
            c.fill = nfill(PALE_GRN); c.font = nfont(bold=True, color=GREEN, size=11)
            c.number_format = "0.00"
        elif col == 9:
            c.fill = nfill(PALE_AMB); c.font = nfont(bold=True, color=AMBER, size=11)
        elif col == 12:
            c.fill = nfill(PALE_GRN); c.font = nfont(bold=True, color=GREEN)
            c.number_format = '€#,##0'
        elif col in [3,4,5,6,7]:
            c.fill = nfill("FFFACD"); c.font = nfont(bold=True, size=11)
        else:
            c.fill = nfill(bg); c.font = nfont()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: INITIATIVE TRACKER
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📋 Initiative Tracker")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:N2")
c = ws4["A1"]
c.value = "INITIATIVE TRACKER  —  One row per AI initiative. Update monthly."
c.fill = nfill(NAVY); c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = nalign("center","center"); ws4.row_dimensions[1].height = 28

header_row(ws4, 3, [
    ("ID",5),("Initiative Name",28),("Track",8),("Owner",14),("Phase",10),
    ("Start Date",12),("Go-Live Date",12),("90-Day Review",12),
    ("Baseline KPI",14),("Target KPI",14),("Current KPI",14),
    ("EUR Saved (actual)",16),("Status",8),("Blockers / Notes",28)
], bg=STEEL, size=9)

sample_inits = [
    ["P-001","Spend Classification Cleanup","B","[Procurement Lead]","Pilot","2026-04-01","2026-06-01","2026-09-01","65% accuracy","85% accuracy","","","🟡 In Progress","Awaiting SAP API access"],
    ["P-002","Contract NLP Scanning","B","[Category Manager]","Planned","2026-04-15","2026-06-15","2026-09-15","4 hrs/contract","45 min/contract","","","🔵 Planned","Tool selection in progress"],
    ["P-003","Supplier Monitoring Agent","D","[Risk Manager]","Planned","2026-05-01","2026-07-01","2026-10-01","Annual review","Continuous","","","🔵 Planned","Jaggaer API mapping needed"],
]
for i, row_data in enumerate(sample_inits):
    row = 4 + i
    bg = row_colors2[i%2]
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=row, column=col, value=val)
        c.fill = nfill(bg); c.font = nfont(size=9 if col==14 else 10)
        c.border = nborder()
        c.alignment = nalign("center" if col not in [2,9,10,11,14] else "left", "center", wrap=(col==14))

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: HARDWARE / COST ESTIMATOR
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🖥️ Hardware & Costs")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:I2")
c = ws5["A1"]
c.value = "ON-PREMISES AI HARDWARE & TOTAL COST ESTIMATOR"
c.fill = nfill(NAVY); c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
c.alignment = nalign("center","center"); ws5.row_dimensions[1].height = 28

section_title(ws5, 4, "  TIER 1: ON-PREMISES INFERENCE SERVER (CONFIDENTIAL data)", cols=9)
header_row(ws5, 5, [("Component",30),("Specification",30),("Qty",6),
                     ("Unit Cost (EUR)",14),("Total Cost (EUR)",14),
                     ("Annual Maintenance",16),("Notes",30),(None,0),(None,0)], bg=STEEL)

hardware = [
    ("GPU Server",              "NVIDIA DGX H100 (8×H100 80GB) or Dell PowerEdge R760xa",  1, 150000, "=C6*D6",  "=E6*0.15", "Runs Llama 3.3 70B / Mistral 3 comfortably"),
    ("GPU (alternative)",       "2× NVIDIA A100 80GB PCIe in standard server (budget path)", 1, 45000,  "=C7*D7",  "=E7*0.15", "Sufficient for most procurement tasks"),
    ("CPU Server (orchestration)","Dell PowerEdge R650 or HP ProLiant DL380",               2, 8000,   "=C8*D8",  "=E8*0.15", "Runs OpenClaw, LangChain, vector DB"),
    ("NVMe Storage",            "4TB NVMe SSD array for model weights + vector DB",          2, 3000,   "=C9*D9",  "=E9*0.15", "Llama 3.3 70B = ~140GB; embeddings storage"),
    ("Network Switch",          "10GbE managed switch for VLAN isolation",                  1, 2500,   "=C10*D10","=E10*0.15","Dedicated VLAN for AI inference layer"),
    ("UPS",                     "APC Smart-UPS 3000VA",                                     2, 1800,   "=C11*D11","=E11*0.15","Power protection for inference servers"),
    ("Rack",                    "42U server rack with PDU and cable management",             1, 1200,   "=C12*D12","=E12*0.15",""),
]

for i, row_data in enumerate(hardware):
    row = 6 + i
    bg = row_colors2[i%2]
    data_cell(ws5, row, 1, row_data[0], bg=bg, bold=True)
    data_cell(ws5, row, 2, row_data[1], bg=bg)
    data_cell(ws5, row, 3, row_data[2], bg=bg, align="center")
    c = ws5.cell(row=row, column=4, value=row_data[3])
    c.fill = nfill("FFFACD"); c.font = nfont(bold=True); c.border = nborder(); c.alignment = nalign("center"); c.number_format = '€#,##0'
    for col_idx, formula in [(5, row_data[4]), (6, row_data[5])]:
        c2 = ws5.cell(row=row, column=col_idx, value=formula)
        c2.fill = nfill(PALE_GRN); c2.font = nfont(color=GREEN, bold=True); c2.border = nborder()
        c2.alignment = nalign("center"); c2.number_format = '€#,##0'
    data_cell(ws5, row, 7, row_data[6], bg=bg, fg=MID_GREY)

# Totals
tot_hw_row = 13
ws5.merge_cells(f"A{tot_hw_row}:D{tot_hw_row}")
c = ws5.cell(row=tot_hw_row, column=1, value="HARDWARE TOTAL (CAPEX)")
c.fill = nfill(NAVY); c.font = nfont(bold=True, color=WHITE, size=12); c.border = nborder(); c.alignment = nalign("right")
for col, formula in [(5,"=SUM(E6:E12)"), (6,"=SUM(F6:F12)")]:
    c = ws5.cell(row=tot_hw_row, column=col, value=formula)
    c.fill = nfill(NAVY); c.font = nfont(bold=True, color=WHITE, size=12); c.border = nborder()
    c.alignment = nalign("center"); c.number_format = '€#,##0'

section_title(ws5, 15, "  SOFTWARE LICENCES & RUNNING COSTS (ANNUAL OPEX)", cols=9)
header_row(ws5, 16, [("Software / Service",30),("Description",35),("Licence Model",14),
                      ("Annual Cost (EUR)",14),("Notes",30),(None,0),(None,0),(None,0),(None,0)], bg=STEEL)

software = [
    ("Ollama / vLLM",        "LLM inference runtime for on-prem models",          "Open Source",       0,         "Free — self-hosted"),
    ("Llama 3.3 / Mistral",  "Foundation LLM models",                             "Open Source (Meta/Mistral)", 0, "Free weights — no per-token cost"),
    ("OpenClaw",             "Agent orchestration platform",                       "npm install",       0,         "Open source — self-hosted"),
    ("ChromaDB / Qdrant",    "Vector database for RAG / document retrieval",       "Open Source",       0,         "Self-hosted on CPU server"),
    ("LangChain",            "Agent workflow and RAG framework",                   "Open Source",       0,         "Python library"),
    ("ELK Stack",            "Elasticsearch + Logstash + Kibana for audit logs",  "Open Source / Cloud", 3600,    "€300/month Elastic Cloud basic tier"),
    ("Grafana",              "Real-time monitoring dashboard",                     "Open Source",       0,         "Self-hosted"),
    ("Docker / Kubernetes",  "Container orchestration for agent sandboxing",       "Open Source",       0,         "Rancher Desktop or K3s"),
    ("Azure OpenAI (Tier 2)","Private cloud LLM for INTERNAL data (EU region)",   "Pay-per-token",     24000,     "~€2K/month for moderate usage"),
    ("ElevenLabs",           "Voice AI for conversational agent interface",        "API subscription",  1800,      "Professional plan ~€150/month"),
    ("Power BI / Tableau",   "Dashboard and reporting for value measurement",      "Existing licence",  0,         "Use existing Statkraft licence"),
    ("Maintenance / IT FTE", "0.5 FTE internal IT support for AI infrastructure",  "Internal",         60000,     "0.5 × €120K fully loaded FTE"),
]

for i, row_data in enumerate(software):
    row = 17 + i
    bg = row_colors2[i%2]
    data_cell(ws5, row, 1, row_data[0], bg=bg, bold=True)
    data_cell(ws5, row, 2, row_data[1], bg=bg)
    c_lic = ws5.cell(row=row, column=3, value=row_data[2])
    lic_bg = PALE_GRN if row_data[2]=="Open Source" or row_data[2].startswith("Open Source") else "FFFACD"
    c_lic.fill = nfill(lic_bg); c_lic.font = nfont(color=GREEN if lic_bg==PALE_GRN else AMBER)
    c_lic.border = nborder(); c_lic.alignment = nalign("center")
    c_cost = ws5.cell(row=row, column=4, value=row_data[3])
    c_cost.fill = nfill(PALE_GRN if row_data[3]==0 else PALE_AMB)
    c_cost.font = nfont(bold=True, color=GREEN if row_data[3]==0 else AMBER)
    c_cost.border = nborder(); c_cost.alignment = nalign("center"); c_cost.number_format = '€#,##0'
    data_cell(ws5, row, 5, row_data[4], bg=bg, fg=MID_GREY)

# Software total
tot_sw_row = 29
ws5.merge_cells(f"A{tot_sw_row}:C{tot_sw_row}")
c = ws5.cell(row=tot_sw_row, column=1, value="ANNUAL SOFTWARE / OPEX TOTAL")
c.fill = nfill(NAVY); c.font = nfont(bold=True, color=WHITE, size=12); c.border = nborder(); c.alignment = nalign("right")
c2 = ws5.cell(row=tot_sw_row, column=4, value=f"=SUM(D17:D28)")
c2.fill = nfill(NAVY); c2.font = nfont(bold=True, color=WHITE, size=12)
c2.border = nborder(); c2.alignment = nalign("center"); c2.number_format = '€#,##0'

section_title(ws5, 31, "  TOTAL COST OF OWNERSHIP SUMMARY", cols=9, bg=NAVY)
tco_rows = [
    ("Year 1 CAPEX (hardware)",    "=E13",  "One-time hardware investment"),
    ("Year 1 OPEX (software + IT)","=D29",  "Annual running costs"),
    ("Year 1 TOTAL INVESTMENT",    "=E32+E33", ""),
    ("Year 2+ annual cost",        "=D29+F13*0.15", "Maintenance + OPEX only"),
    ("3-Year Total Cost of Ownership", "=E34+(E35*2)", ""),
    ("Estimated Annual Value (from Value Calculator)", "='💰 Value Calculator'!E26", ""),
    ("3-Year Net ROI",             "='💰 Value Calculator'!E26*3-E36", ""),
]
for i, (label, formula, note) in enumerate(tco_rows):
    row = 32 + i
    c1 = ws5.cell(row=row, column=1, value=label); c1.fill = nfill(LTBLUE if i%2==0 else WHITE)
    c1.font = nfont(bold=(i in [2,4,6]), size=11); c1.border = nborder(); ws5.merge_cells(f"A{row}:D{row}"); c1.alignment = nalign("right")
    c2 = ws5.cell(row=row, column=5, value=formula)
    c2.fill = nfill(PALE_GRN if i in [5,6] else (NAVY if i in [2,4] else LTBLUE if i%2==0 else WHITE))
    c2.font = nfont(bold=True, color=WHITE if i in [2,4] else GREEN, size=12 if i in [2,4,6] else 11)
    c2.border = nborder(); c2.alignment = nalign("center"); c2.number_format = '€#,##0'
    if note:
        data_cell(ws5, row, 6, note, bg=c1.fill.fgColor.rgb, fg=MID_GREY)

# Set column widths for ws5
for col, width in [(1,32),(2,36),(3,16),(4,16),(5,16),(6,16),(7,32)]:
    set_col_width(ws5, col, width)

# ── FINAL FORMATTING ──────────────────────────────────────────────────────────
# Freeze panes on all sheets
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.freeze_panes = "A8" if ws in [ws1] else ("A4" if ws in [ws4] else "A8")

wb.save("/Users/jonathonmilne/.openclaw/workspace/venture/Statkraft_Procurement_AI_Value_Model.xlsx")
print("Excel saved.")
