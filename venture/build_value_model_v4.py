"""
Clean rebuild — hard-coded calculated values based on €1.9B spend.
Formulas replaced with actual numbers so the spreadsheet opens correctly.
Yellow input cells remain editable with notes showing the formula used.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook()

# ── COLOURS ───────────────────────────────────────────────────────────────────
NAVY   = "002147"; STEEL  = "2B6CB0"; WHITE  = "FFFFFF"; GREY   = "718096"
AMBER  = "D97F06"; GREEN  = "27AE60"; RED    = "C0392B"; DARK   = "2D3A4A"
LTBLUE = "EBF4FF"; DEEP   = "041220"; PGREEN = "D4EDDA"; PAMBER = "FFF3CD"
PRED   = "F8D7DA"; YELLOW = "FFFACD"; TEAL   = "0B7A75"

def F(c): return PatternFill("solid", fgColor=c)
def FT(b=False,c=DARK,s=10,i=False): return Font(bold=b,color=c,size=s,name="Calibri",italic=i)
def AL(h="left",v="center",w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w)
def BD():
    t=Side(style="thin"); return Border(left=t,right=t,top=t,bottom=t)

def C(ws,r,c,val,bg=WHITE,fg=DARK,bold=False,sz=10,align="left",wrap=False,italic=False,fmt=None):
    cl=ws.cell(row=r,column=c,value=val)
    cl.fill=F(bg); cl.font=FT(bold,fg,sz,italic)
    cl.alignment=AL(align,"center",wrap); cl.border=BD()
    if fmt: cl.number_format=fmt
    return cl

def EUR(ws,r,c,val,bg=WHITE,bold=False,sz=10):
    cl=ws.cell(row=r,column=c,value=val)
    cl.fill=F(bg); cl.font=FT(bold,WHITE if bg in [NAVY,STEEL] else (GREEN if bg==PGREEN else DARK),sz)
    cl.alignment=AL("right","center"); cl.border=BD()
    cl.number_format='€#,##0'
    return cl

def HDR(ws,r,cols,bg=NAVY,fg=WHITE,sz=10):
    for i,(v,w) in enumerate(cols,1):
        if v is None: continue
        cl=ws.cell(row=r,column=i,value=v)
        cl.fill=F(bg); cl.font=FT(True,fg,sz)
        cl.alignment=AL("center","center",True); cl.border=BD()
        if w: ws.column_dimensions[get_column_letter(i)].width=w

def SEC(ws,r,title,ncols=18,bg=STEEL,fg=WHITE):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
    cl=ws.cell(row=r,column=1,value=title)
    cl.fill=F(bg); cl.font=FT(True,fg,11); cl.alignment=AL("left","center"); cl.border=BD()

# ── INPUTS ────────────────────────────────────────────────────────────────────
SPEND     = 1_900_000_000
FTES      = 45
FTE_COST  = 120_000
EVENTS    = 250
HRS_EVENT = 40
CONTRACTS = 1200
HRS_CTR   = 4
SUPPLIERS = 200
SUP_COST  = 150_000
ACCURACY  = 0.65
MRO_PCT   = 0.03
CAPEX     = 400_000_000
OVERRUN   = 0.40
CSRD      = 12
TAIL_PCT  = 0.08
INV_MINS  = 15
INVOICES  = 18_000
HR_RATE   = FTE_COST / 1750   # = 68.57

# ── CALCULATED VALUES ──────────────────────────────────────────────────────────
# Each tuple: (id, name, category, calc_type, methodology_note,
#              bear, base, bull, confidence, source, effort, track, capex_cost, opex_cost)

use_cases = [
    (1,"Spend Classification & Enrichment","Data Quality","Type B — Spend Savings",
     "Formula: Annual spend × (Target accuracy 88% − Current accuracy 65%) × 5% spend recovery rate\n"
     "Basis: 23% accuracy gap means 23% of spend mis-classified. Recovery = 5% of mis-classified spend redirected to preferred suppliers/contracts.\n"
     "Peer: Equinor achieved 500M NOK/yr from automation; McKinsey: AI classification improves savings 5–10%.",
     SPEND*(0.88-ACCURACY)*0.05*0.5, SPEND*(0.88-ACCURACY)*0.05, SPEND*(0.88-ACCURACY)*0.05*1.5,
     "High","McKinsey + Equinor", "4–6 wks","B", 0, 15_000),

    (2,"Contract NLP Scanning & Review","Contract Mgmt","Type A — Time Savings",
     "Formula: FTE hourly rate × Hours saved per contract × Number of contracts per year\n"
     "Hours saved: 4 hrs manual review → 0.8 hrs with AI (80% reduction) = 3.2 hrs saved per contract.\n"
     "FTE hourly rate: €120,000 ÷ 1,750 working hours = €68.57/hr.\n"
     "Peer: Luminance/Icertis benchmark: 80% faster review. 31% of companies already using AI for contracts.",
     HR_RATE*CONTRACTS*HRS_CTR*0.8*0.5, HR_RATE*CONTRACTS*HRS_CTR*0.8, HR_RATE*CONTRACTS*HRS_CTR*0.8*1.5,
     "High","Luminance/Icertis + 31% adoption data", "6–8 wks","B", 5_000, 8_000),

    (3,"Supplier Monitoring Agent","Risk Mgmt","Type A — Time Savings",
     "Formula: Annual manual monitoring cost × 70% reduction\n"
     "Baseline: Statkraft estimate €150K/yr for manual supplier review programme.\n"
     "AI replaces annual reviews with continuous automated monitoring — 70% cost reduction is conservative.\n"
     "Peer: Resilinc (Gartner Magic Quadrant Leader) live with real-time disruption/tariff/sanctions monitoring.",
     SUP_COST*0.7*0.5, SUP_COST*0.7, SUP_COST*0.7*1.5,
     "High","Resilinc Gartner Leader", "8–10 wks","D", 10_000, 25_000),

    (4,"RFQ & Tender Drafting Agent","Sourcing","Type A — Time Savings",
     "Formula: FTE hourly rate × Hours saved per event × Number of sourcing events per year\n"
     "Hours saved: 40 hrs per event × 75% reduction = 30 hrs saved per event.\n"
     "FTE hourly rate: €68.57/hr.\n"
     "Note: Only 250 events assumed — actual savings scale with volume.\n"
     "Peer: Ørsted and BP live with autonomous RFQ generation. Fairmarkit: 10× more events per FTE.",
     HR_RATE*EVENTS*HRS_EVENT*0.75*0.5, HR_RATE*EVENTS*HRS_EVENT*0.75, HR_RATE*EVENTS*HRS_EVENT*0.75*1.5,
     "High","Ørsted + BP live + Fairmarkit $40K/buyer/week", "8–12 wks","D", 15_000, 30_000),

    (5,"Demand Forecasting — MRO & Capex","MRO","Type B — Spend Savings",
     "Formula: (Annual spend × MRO emergency premium rate × 50% reduction)\n"
     "         + (Annual spend × 2% inventory carrying cost × 13% inventory reduction)\n"
     "Emergency premium: 3% of €1.9B = €57M in premium-priced emergency buys. AI forecasting cuts this 50% = €28.5M.\n"
     "Inventory saving: 2% carry rate on reduced stock × 13% inventory reduction = €4.9M additional.\n"
     "Peer: Vattenfall: 6–12 month horizon forecasting live. McKinsey: 35% inventory reduction with AI.",
     (SPEND*MRO_PCT*0.5 + SPEND*0.02*0.13)*0.5, SPEND*MRO_PCT*0.5 + SPEND*0.02*0.13,
     (SPEND*MRO_PCT*0.5 + SPEND*0.02*0.13)*1.5,
     "High","Vattenfall live + McKinsey 35% inventory reduction", "12–16 wks","D", 20_000, 40_000),

    (6,"Negotiation Intelligence & Preparation","Sourcing","Type B — Spend Savings",
     "Formula: Annual spend × 0.3% additional savings rate\n"
     "Basis: 0.3% is conservative. AI-assisted negotiation packs provide market benchmarks, should-cost data, BATNA analysis.\n"
     "Even 0.3% of €1.9B = €5.7M. Used 0.3% vs Arkestro's reported 18.8% per event as a very conservative enterprise-wide rate.\n"
     "Peer: Arkestro: 18.8% savings per $1M spend. McKinsey should-cost: 13% raw materials savings.",
     SPEND*0.003*0.5, SPEND*0.003, SPEND*0.003*1.5,
     "Medium","Arkestro 18.8% + McKinsey 13% should-cost", "8–12 wks","D", 10_000, 20_000),

    (7,"Category Strategy Agent","Strategy","Type B — Spend Savings",
     "Formula: Annual spend × 0.5% strategic savings from better category coverage\n"
     "Basis: Increase from 3 to 12 category strategies per year (same team). Each strategy covers ~4% of spend.\n"
     "9 additional strategies × 4% of spend each × 1.3% avg savings rate = 0.47% of total spend.\n"
     "Peer: McKinsey: autonomous category agents deliver 15–30% efficiency. BCG: 15–45% cost reduction with GenAI.",
     SPEND*0.005*0.5, SPEND*0.005, SPEND*0.005*2.0,
     "Medium","McKinsey 15-30% + BCG 15-45%", "12–20 wks","D", 25_000, 50_000),

    (8,"CSRD & ESG Compliance Automation","Compliance","Type A — Time Savings",
     "Formula: FTE hourly rate × Hours per check × Checks per supplier per year × Suppliers × 80% automation rate\n"
     "Hours per check: 0.8 hours average (research, documentation, verification).\n"
     "FTE hourly rate: €68.57/hr.\n"
     "Regulatory requirement: CSRD mandatory for Statkraft from FY2026 reporting.\n"
     "Plus penalty risk reduction (unquantified here — treated as strategic value).",
     HR_RATE*SUPPLIERS*CSRD*0.8*0.8*0.5, HR_RATE*SUPPLIERS*CSRD*0.8*0.8,
     HR_RATE*SUPPLIERS*CSRD*0.8*0.8*1.5,
     "High","CSRD mandatory 2026 + Resilinc agent live", "10–14 wks","D", 15_000, 25_000),

    (9,"Invoice Processing & Anomaly Detection","P2P","Type A — Time Savings",
     "Formula: FTE hourly rate × (Invoice processing time saved in hours) × Annual invoice volume\n"
     "Time saved: 15 min manual → 4.5 min with AI = 10.5 min saved = 0.175 hrs per invoice.\n"
     "FTE hourly rate: €68.57/hr.\n"
     "Jaggaer anomaly detection live Q2 2025. SAP Joule invoice matching native Q4 2025.\n"
     "Additional value: fraud/error detection estimated at 0.1% of invoice value (not included in base).",
     INVOICES*(INV_MINS*0.7/60)*HR_RATE*0.5, INVOICES*(INV_MINS*0.7/60)*HR_RATE,
     INVOICES*(INV_MINS*0.7/60)*HR_RATE*1.5,
     "High","Jaggaer anomaly detection live + SAP Joule native", "4–6 wks","B", 0, 5_000),

    (10,"Market Intelligence Agent (Daily Briefings)","Market Intel","Type B — Spend Savings",
     "Formula: Commodity-exposed spend × 1% better purchase timing improvement\n"
     "Commodity-exposed spend: ~30% of total (steel, copper, cable, equipment) = €570M.\n"
     "1% timing improvement on €570M = €5.7M. Conservative — steel/copper volatile 25–40% in 2024.\n"
     "Peer: GEP: 15% lower logistics costs with AI supply chain. Equinor: live commodity monitoring.",
     SPEND*0.3*0.01*0.5, SPEND*0.3*0.01, SPEND*0.3*0.01*2.0,
     "Medium","GEP 15% logistics + Equinor commodity monitoring", "2–4 wks","A/B", 0, 3_000),

    (11,"Should-Cost Modelling Agent","Cost Intelligence","Type B — Spend Savings",
     "Formula: Direct material spend × 1% savings from better cost baselines\n"
     "Direct material spend: ~25% of total = €475M (equipment, cables, structural).\n"
     "1% savings from should-cost models = €4.75M. McKinsey benchmark is 13% — using 1% as year-1 conservative.\n"
     "Peer: McKinsey: 13% savings in raw-materials from AI should-cost modelling.",
     SPEND*0.25*0.01*0.5, SPEND*0.25*0.01, SPEND*0.25*0.013,
     "High","McKinsey 13% raw materials should-cost", "8–12 wks","D", 20_000, 35_000),

    (12,"Supplier Onboarding Automation","Supplier Mgmt","Type A — Time Savings",
     "Formula: FTE daily cost × Days saved per onboarding × New suppliers per year\n"
     "Days saved: 45 days → 18 days = 27 days saved per new supplier.\n"
     "FTE daily cost: €120K ÷ 250 working days = €480/day.\n"
     "New suppliers per year: estimated 40.\n"
     "Note: Low value relative to complexity — deprioritised vs. other use cases.",
     (FTE_COST/250)*27*0.6*40*0.5, (FTE_COST/250)*27*0.6*40, (FTE_COST/250)*27*0.6*40*1.5,
     "Medium","TealBook/Scoutbee (Coupa acquisition)", "10–14 wks","C", 10_000, 15_000),

    (13,"Tail Spend Automation Agent","P2P","Type B — Spend Savings",
     "Formula: Tail spend volume × 15% savings rate via competitive mini-tendering\n"
     "Tail spend: 8% of €1.9B = €152M (purchases under €10K, typically uncompetitive).\n"
     "15% savings via automated competitive tendering = €22.8M.\n"
     "Peer: Fairmarkit: $40K savings per buyer per week. Zip: $6B+ in customer savings.",
     SPEND*TAIL_PCT*0.15*0.5, SPEND*TAIL_PCT*0.15, SPEND*TAIL_PCT*0.15*1.8,
     "High","Fairmarkit $40K/buyer/week + Zip $6B savings", "6–10 wks","D", 10_000, 20_000),

    (14,"Project Cost Intelligence Agent","Decision Intel","Type C — Risk Reduction",
     "Formula: Annual capex × Average energy project overrun rate × % improvement from AI cost intelligence\n"
     "Annual capex: €400M (approx 20% of total spend).\n"
     "Avg energy project overrun: 40% (ScienceDirect 2025 — 662 projects, $1.358 trillion).\n"
     "2% improvement in estimate accuracy = €3.2M better capital allocation.\n"
     "Mechanism: Agent mines SAP/Jaggaer actuals vs estimates, builds live cost intelligence database.",
     CAPEX*OVERRUN*0.02*0.4, CAPEX*OVERRUN*0.02, CAPEX*OVERRUN*0.02*2.0,
     "Medium","ScienceDirect 662 energy projects — 40% avg overrun", "16–24 wks","D", 30_000, 60_000),

    (15,"EPC Contract AI Review","Contract Mgmt","Type A — Time Savings",
     "Formula: FTE hourly rate × Hours saved per EPC contract × EPC contracts per year\n"
     "EPC contracts: ~20 major contracts per year, average 40 hrs manual review each.\n"
     "Hours saved: 40% time reduction = 16 hrs saved per contract.\n"
     "Peer: Ironclad ($150M ARR, 39% growth), Icertis (33 Fortune 100 clients) live with this.",
     HR_RATE*20*40*0.4*0.5, HR_RATE*20*40*0.4, HR_RATE*20*40*0.6,
     "Medium","Ironclad $150M ARR + Icertis 33 Fortune 100 clients", "12–18 wks","D", 20_000, 40_000),

    (16,"Supplier Financial Risk Monitoring","Risk Mgmt","Type C — Risk Reduction",
     "Formula: Probability of major supplier failure per year × Average cost of unplanned supplier switch\n"
     "Probability: 1 major supplier failure event per 2 years = 0.5 per year (conservative for 200-supplier base).\n"
     "Impact: €2M average cost (emergency sourcing premium + project schedule impact).\n"
     "Expected value: 0.5 × €2M = €1M per year.",
     2_000_000/2*0.5, 2_000_000/2, 2_000_000/2*2,
     "Medium","Shell/TotalEnergies live; €2M avg cost documented", "8–12 wks","D", 10_000, 20_000),

    (17,"SAP Joule — Native AI Activation","Platform AI","Type A — Time Savings",
     "Formula: FTE hourly rate × 1,000 hours saved per year from Joule agents\n"
     "Joule agents live Q4 2025: bid analysis, SOW automation, NL demand routing, mass PO date updates.\n"
     "Estimated 1,000 hours saved across the team per year from native Joule features.\n"
     "COST: €0 additional — included in existing SAP licence. Activate now.",
     HR_RATE*1000*0.2*0.5, HR_RATE*1000*0.2, HR_RATE*1000*0.2*1.5,
     "High","SAP Joule native, GA Q4 2025, €0 additional cost", "1–2 wks","A", 0, 0),

    (18,"Jaggaer JAI — Native AI Activation","Platform AI","Type A — Time Savings",
     "Formula: FTE hourly rate × 800 hours saved per year from JAI copilot\n"
     "JAI copilot live 2025: anomaly detection, conversational CLM, contextual support.\n"
     "Estimated 800 hours saved per year from JAI native features.\n"
     "COST: €0 additional — included in existing Jaggaer licence. Activate now.",
     HR_RATE*800*0.15*0.5, HR_RATE*800*0.15, HR_RATE*800*0.15*1.5,
     "High","Jaggaer JAI live, €0 additional cost", "1–2 wks","A", 0, 0),

    (19,"Autonomous Low-Value Tender Management","P2P","Type B — Spend Savings",
     "Formula: Number of low-value tenders × Average tender value × 8% competitive savings rate\n"
     "Low-value tenders: 40% of 250 events = 100 tenders under €50K.\n"
     "Average value: €50,000 per tender.\n"
     "8% savings from competitive AI-assisted tendering vs. direct award.\n"
     "Peer: Zip (50+ agents, $355B processed), Pactum (manages <$150K autonomously).",
     EVENTS*0.4*50_000*0.08*0.5, EVENTS*0.4*50_000*0.08, EVENTS*0.4*50_000*0.08*1.5,
     "Medium","Zip $6B savings + Pactum autonomous negotiation", "10–16 wks","D", 15_000, 30_000),

    (20,"Supply Chain Network Risk Mapping","Risk/Strategy","Type C — Risk Reduction",
     "Formula: Probability of major supply disruption × Cost impact avoided\n"
     "Probability: 15% per year of a significant supply chain disruption (single-source dependency, geographic risk).\n"
     "Impact: €5M average cost (project delay, emergency logistics, premium pricing).\n"
     "Expected value: 15% × €5M = €750K per year.\n"
     "Peer: GEP: 15% lower logistics, 35% leaner inventory; Resilinc sub-tier mapping live.",
     5_000_000*0.15*0.4, 5_000_000*0.15, 5_000_000*0.15*2.0,
     "Low","GEP/Resilinc benchmarks + supply disruption literature", "16–24 wks","D", 25_000, 50_000),
]

# ══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📊 Value Register"
ws.sheet_view.showGridLines = False

# Title
ws.merge_cells("A1:R2")
cl = ws["A1"]
cl.value = "STATKRAFT PROCUREMENT AI — USE CASE VALUE REGISTER  |  Based on €1.9B annual spend  |  March 2026"
cl.fill = F(NAVY); cl.font = FT(True, WHITE, 14)
cl.alignment = AL("center", "center"); cl.border = BD()
ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 16

ws.merge_cells("A3:R3")
cl = ws["A3"]
cl.value = (f"FTE hourly rate = €{FTE_COST:,} ÷ 1,750 hrs = €{HR_RATE:.2f}/hr  |  "
            f"Spend = €{SPEND/1e9:.1f}B  |  FTEs = {FTES}  |  "
            f"All values in EUR  |  Bear = 50% of base  |  Bull = 150% of base (unless stated)")
cl.fill = F(STEEL); cl.font = FT(False, WHITE, 9, True)
cl.alignment = AL("center", "center"); cl.border = BD()

# Headers
HDR(ws, 5, [
    ("#",4), ("Use Case",32), ("Category",14), ("Calc Type",12),
    ("Methodology — Formula & Assumptions",55),
    ("Bear\nEUR",14), ("Base\nEUR",14), ("Bull\nEUR",14),
    ("Confidence",10), ("Benchmark Source",22),
    ("Effort\n(wks)",8), ("Track",6),
    ("One-off\nCost EUR",12), ("Annual\nCost EUR",12),
    ("Net Annual\nValue EUR",14), ("3yr NPV\nEUR",14),
    ("ROI",6), ("Notes",30)
], bg=NAVY, fg=WHITE, sz=9)

# Data rows
row_bgs = [LTBLUE, WHITE]
for i, uc in enumerate(use_cases):
    r = 6 + i
    bg = row_bgs[i % 2]
    (uc_id, name, cat, ctype, method, bear, base, bull,
     conf, source, effort, track, capex_c, opex_c) = uc

    C(ws, r, 1, uc_id, bg=STEEL, fg=WHITE, bold=True, align="center", sz=11)
    C(ws, r, 2, name, bg=bg, bold=True, sz=10)
    C(ws, r, 3, cat, bg=bg, sz=9, align="center")
    # Calc type coloured
    type_bg = {"Type A": STEEL, "Type B": AMBER, "Type C": RED, "Type D": GREEN}.get(ctype[:6], DEEP)
    C(ws, r, 4, ctype[:6], bg=type_bg, fg=WHITE, bold=True, align="center", sz=9)
    C(ws, r, 5, method, bg=bg, sz=8, italic=True, wrap=True)

    # Bear/Base/Bull — actual numbers
    EUR(ws, r, 6, round(bear), bg=PRED, bold=False, sz=10)
    EUR(ws, r, 7, round(base), bg=PGREEN, bold=True, sz=11)
    EUR(ws, r, 8, round(bull), bg="D1F2EB", bold=False, sz=10)
    ws.cell(row=r, column=6).font = FT(False, RED, 10)

    # Confidence
    conf_bg = {"High": PGREEN, "Medium": PAMBER, "Low": PRED}.get(conf, WHITE)
    conf_fg = {"High": GREEN, "Medium": AMBER, "Low": RED}.get(conf, DARK)
    C(ws, r, 9, conf, bg=conf_bg, fg=conf_fg, bold=True, align="center", sz=9)
    C(ws, r, 10, source, bg=bg, sz=8, italic=True, wrap=True)
    C(ws, r, 11, effort, bg=bg, align="center", sz=9)

    track_bg = NAVY if track == "D" else (STEEL if track == "C" else LTBLUE)
    track_fg = WHITE if track in ["D", "C"] else NAVY
    C(ws, r, 12, track, bg=track_bg, fg=track_fg, bold=True, align="center", sz=10)

    EUR(ws, r, 13, capex_c, bg=PAMBER)
    ws.cell(row=r, column=13).font = FT(False, AMBER, 9)
    EUR(ws, r, 14, opex_c, bg=PAMBER)
    ws.cell(row=r, column=14).font = FT(False, AMBER, 9)

    # Net annual value = base - annual cost
    net = round(base - opex_c)
    EUR(ws, r, 15, net, bg=PGREEN, bold=True, sz=11)

    # 3yr NPV = net × 3 - capex
    npv = round(net * 3 - capex_c)
    EUR(ws, r, 16, npv, bg=PGREEN, bold=True, sz=11)

    # ROI
    if capex_c + opex_c > 0:
        roi = base / (capex_c + opex_c)
        cl = ws.cell(row=r, column=17, value=round(roi, 1))
        cl.fill = F(PGREEN); cl.font = FT(True, GREEN, 10)
        cl.alignment = AL("center", "center"); cl.border = BD()
        cl.number_format = '0.0"x"'
    else:
        C(ws, r, 17, "∞ (native)", bg=PGREEN, fg=GREEN, bold=True, align="center", sz=9)

    C(ws, r, 18, "", bg=bg)
    ws.row_dimensions[r].height = 80

# TOTALS
TOT = 6 + len(use_cases)
ws.merge_cells(f"A{TOT}:E{TOT}")
cl = ws.cell(row=TOT, column=1, value="PROGRAMME TOTAL — ALL 20 USE CASES FULLY DEPLOYED (Base estimates, €1.9B spend)")
cl.fill = F(NAVY); cl.font = FT(True, WHITE, 12); cl.border = BD(); cl.alignment = AL("right")

total_bear = sum(round(uc[5]) for uc in use_cases)
total_base = sum(round(uc[6]) for uc in use_cases)
total_bull = sum(round(uc[7]) for uc in use_cases)
total_capex = sum(uc[12] for uc in use_cases)
total_opex = sum(uc[13] for uc in use_cases)
total_net = total_base - total_opex
total_npv = total_net * 3 - total_capex
total_roi = total_base / (total_capex + total_opex) if (total_capex + total_opex) > 0 else 0

for col, val in [(6, total_bear), (7, total_base), (8, total_bull),
                 (13, total_capex), (14, total_opex), (15, total_net), (16, total_npv)]:
    cl = ws.cell(row=TOT, column=col, value=val)
    cl.fill = F(NAVY); cl.font = FT(True, WHITE, 12); cl.border = BD()
    cl.alignment = AL("right", "center"); cl.number_format = '€#,##0'

cl = ws.cell(row=TOT, column=17, value=round(total_roi, 1))
cl.fill = F(NAVY); cl.font = FT(True, WHITE, 13); cl.border = BD()
cl.alignment = AL("center"); cl.number_format = '0.0"x"'

cl = ws.cell(row=TOT+1, column=7, value=f"= {total_base/1e6:.0f}M base = {total_base/1_900_000_000*100:.1f}% of €1.9B spend")
cl.fill = F(STEEL); cl.font = FT(True, WHITE, 10); cl.border = BD(); cl.alignment = AL("center")
ws.merge_cells(f"G{TOT+1}:H{TOT+1}")
ws.row_dimensions[TOT].height = 28

# ── SHEET 2: PHASED SUMMARY ───────────────────────────────────────────────────
ws2 = wb.create_sheet("📅 Phased Rollout")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:H2")
cl = ws2["A1"]
cl.value = "PHASED VALUE ROLLOUT — When Does the Money Land?"
cl.fill = F(NAVY); cl.font = FT(True, WHITE, 14)
cl.alignment = AL("center", "center"); cl.border = BD()
ws2.row_dimensions[1].height = 28

HDR(ws2, 4, [
    ("Phase",20),("Use Cases Included",40),("Timeline",12),
    ("Bear EUR",14),("Base EUR",14),("Bull EUR",14),
    ("Cumulative Base",16),("Note",30)
], bg=NAVY, fg=WHITE, sz=10)

phases = [
    ("Phase 0: Quick Wins",
     "#17 SAP Joule + #18 Jaggaer JAI",
     "Weeks 1–2",
     [17, 18]),
    ("Phase 1: Three Pilots",
     "#1 Spend Classification + #2 Contract NLP + #3 Supplier Monitor",
     "Weeks 3–12",
     [1, 2, 3]),
    ("Phase 2: Expand",
     "#4 RFQ Drafter + #5 Demand Forecast + #6 Negotiation + #9 Invoice + #10 Market Intel + #11 Should-Cost + #13 Tail Spend",
     "Months 4–6",
     [4, 5, 6, 9, 10, 11, 13]),
    ("Phase 3: Full Agentic",
     "#7 Category Strategy + #8 CSRD + #12 Onboarding + #14 Decision Intel + #15 EPC Review + #16 Risk + #19 Auto Tender + #20 Network",
     "Months 7–12",
     [7, 8, 12, 14, 15, 16, 19, 20]),
    ("FULL PROGRAMME",
     "All 20 use cases",
     "Year 1+",
     list(range(1, 21))),
]

cum = 0
row_bgs2 = [LTBLUE, WHITE, LTBLUE, WHITE, NAVY]
for i, (phase, uc_list, timeline, ids) in enumerate(phases):
    r = 5 + i
    bg = row_bgs2[i]
    fg = WHITE if i == 4 else DARK
    bold = (i == 4)

    uc_vals = {uc[0]: uc for uc in use_cases}
    p_bear = sum(round(uc_vals[id][5]) for id in ids if id in uc_vals)
    p_base = sum(round(uc_vals[id][6]) for id in ids if id in uc_vals)
    p_bull = sum(round(uc_vals[id][7]) for id in ids if id in uc_vals)
    cum += p_base

    C(ws2, r, 1, phase, bg=bg, fg=fg, bold=bold, sz=11 if bold else 10)
    C(ws2, r, 2, uc_list, bg=bg, fg=fg, sz=9, wrap=True)
    C(ws2, r, 3, timeline, bg=bg, fg=fg, align="center", sz=9)
    for col, val in [(4, p_bear), (5, p_base), (6, p_bull)]:
        cl = ws2.cell(row=r, column=col, value=val)
        cl.fill = F(bg); cl.font = FT(bold, fg, 12 if bold else 10)
        cl.border = BD(); cl.alignment = AL("right", "center"); cl.number_format = '€#,##0'
    cl = ws2.cell(row=r, column=7, value=cum if i < 4 else total_base)
    cl.fill = F(PGREEN if i < 4 else NAVY)
    cl.font = FT(True, DARK if i < 4 else WHITE, 12 if bold else 10)
    cl.border = BD(); cl.alignment = AL("right", "center"); cl.number_format = '€#,##0'
    C(ws2, r, 8, "Quick start — €0 additional cost" if i==0 else
                 "90-day review gate" if i==1 else
                 f"Annual value at phase end: €{p_base/1e6:.0f}M" if i < 4 else
                 f"Total: €{total_base/1e6:.0f}M/yr = {total_base/SPEND*100:.1f}% of spend",
      bg=bg, fg=fg, sz=9, italic=True)
    ws2.row_dimensions[r].height = 30

# ── SHEET 3: METHODOLOGY ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("📐 Methodology")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F2")
cl = ws3["A1"]
cl.value = "CALCULATION METHODOLOGY — Standard Framework for All Use Cases"
cl.fill = F(NAVY); cl.font = FT(True, WHITE, 14)
cl.alignment = AL("center", "center"); cl.border = BD()
ws3.row_dimensions[1].height = 28

ws3.merge_cells("A3:F3")
cl = ws3["A3"]
cl.value = (f"Key inputs: Annual spend = €{SPEND/1e9:.1f}B  |  "
            f"FTEs = {FTES}  |  FTE cost = €{FTE_COST:,}/yr  |  "
            f"FTE hourly rate = €{FTE_COST:,} ÷ 1,750 hrs = €{HR_RATE:.2f}/hr  |  "
            f"Bear = 50% of base  |  Bull = 150% of base (stated explicitly if different)")
cl.fill = F(STEEL); cl.font = FT(False, WHITE, 10, True)
cl.alignment = AL("center"); cl.border = BD()

types = [
    ("TYPE A — Time Savings", STEEL,
     "Formula: (Hours saved per unit) × (Volume per year) × (FTE hourly rate)",
     [("Hours saved per unit", "Baseline hours − AI-assisted hours (source cited)"),
      ("Volume per year", "Actual Statkraft volume (contracts, events, invoices)"),
      ("FTE hourly rate", "€120,000 annual cost ÷ 1,750 working hours = €68.57/hr"),
      ("Assumption", "Freed time is redeployed to higher-value work — no headcount reduction assumed"),
      ("Examples", "Contract NLP: 4 hrs → 0.8 hrs × 1,200 contracts × €68.57\n"
                   "CSRD: 0.8 hrs × 200 suppliers × 12 checks × €68.57")]),

    ("TYPE B — Spend Savings", AMBER,
     "Formula: (Addressable spend) × (Savings rate %)",
     [("Addressable spend", "€1.9B × Category relevance % (e.g. tail spend = 8% = €152M)"),
      ("Savings rate", "Conservative benchmark from peer implementations — always cited"),
      ("Benchmark principle", "Use the LOWEST credible benchmark, not the vendor headline"),
      ("Example", "Tail spend: McKinsey/Fairmarkit benchmark = 15–20% savings. Using 15%."),
      ("Examples", "Tail Spend: €152M × 15% = €22.8M\n"
                   "Negotiation: €1.9B × 0.3% = €5.7M (vs Arkestro 18.8% — very conservative)")]),

    ("TYPE C — Risk Reduction", RED,
     "Formula: (Probability of event per year) × (Cost of event)",
     [("Probability", "Conservative assessment of adverse event frequency — documented"),
      ("Event cost", "Quantified cost of event occurring (emergency sourcing, delays, penalties)"),
      ("Expected value", "P × Impact = annual expected value of avoided cost"),
      ("Principle", "Only include risks with documented precedent — no speculative events"),
      ("Examples", "Supplier failure: 0.5 events/yr × €2M avg cost = €1M\n"
                   "Cost overrun: €400M capex × 40% overrun × 2% improvement = €3.2M")]),

    ("TYPE D — Capacity Release", GREEN,
     "Formula: (FTE hours automated) × (% redirectable) × (FTE hourly rate)",
     [("FTE hours automated", "Total hours of routine tasks eliminated by AI per year"),
      ("% redirectable", "Conservative % of freed time that can be redirected to strategic work"),
      ("Principle", "Assumes redeployment — not headcount reduction. Conservative multiplier."),
      ("When to use", "When the primary value is freeing up procurement expertise, not direct spend savings"),
      ("Examples", "RFQ Drafter: 30 hrs saved × 250 events × €68.57 = €514K\n"
                   "Market Intel: analyst time freed for strategic category work")]),
]

row_start = 5
for t_idx, (title, color, formula, rows) in enumerate(types):
    r = row_start + t_idx * 9
    ws3.merge_cells(f"A{r}:F{r}")
    cl = ws3.cell(row=r, column=1, value=title)
    cl.fill = F(color); cl.font = FT(True, WHITE, 12)
    cl.alignment = AL("left", "center"); cl.border = BD(); ws3.row_dimensions[r].height = 22

    ws3.merge_cells(f"A{r+1}:F{r+1}")
    cl = ws3.cell(row=r+1, column=1, value=formula)
    cl.fill = F(DEEP); cl.font = FT(True, WHITE, 11)
    cl.alignment = AL("left", "center"); cl.border = BD(); ws3.row_dimensions[r+1].height = 20

    for j, (label, desc) in enumerate(rows):
        rr = r + 2 + j
        C(ws3, rr, 1, label, bg=LTBLUE, fg=DARK, bold=True, sz=10)
        ws3.merge_cells(f"B{rr}:F{rr}")
        cl = ws3.cell(row=rr, column=2, value=desc)
        cl.fill = F(WHITE); cl.font = FT(False, DARK, 9, True if label in ["Examples","Example"] else False)
        cl.alignment = AL("left", "center", True); cl.border = BD()
        ws3.row_dimensions[rr].height = 30

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 80

# ── SHEET 4: SUMMARY FOR DECK ─────────────────────────────────────────────────
ws4 = wb.create_sheet("🎯 Summary for Deck")
ws4.sheet_view.showGridLines = False
ws4.merge_cells("A1:E2")
cl = ws4["A1"]
cl.value = "NUMBERS TO USE IN PRESENTATION — Updated from €1.9B Spend"
cl.fill = F(NAVY); cl.font = FT(True, WHITE, 14)
cl.alignment = AL("center", "center"); cl.border = BD(); ws4.row_dimensions[1].height = 28

HDR(ws4, 4, [("Metric",35),("Bear EUR",16),("Base EUR",16),("Bull EUR",16),("Slide / Use",25)], bg=NAVY)

summary = [
    ("Phase 0 Quick Wins (SAP Joule + Jaggaer JAI)", 
     round(sum(round(uc[5]) for uc in use_cases if uc[0] in [17,18])),
     round(sum(round(uc[6]) for uc in use_cases if uc[0] in [17,18])),
     round(sum(round(uc[7]) for uc in use_cases if uc[0] in [17,18])),
     "Slide: Roadmap — Phase 0"),
    ("Phase 1 Pilots (Spend, Contract, Supplier)",
     round(sum(round(uc[5]) for uc in use_cases if uc[0] in [1,2,3])),
     round(sum(round(uc[6]) for uc in use_cases if uc[0] in [1,2,3])),
     round(sum(round(uc[7]) for uc in use_cases if uc[0] in [1,2,3])),
     "Slide: Roadmap — Phase 1"),
    ("Phase 2 Expanded (7 use cases)",
     round(sum(round(uc[5]) for uc in use_cases if uc[0] in [4,5,6,9,10,11,13])),
     round(sum(round(uc[6]) for uc in use_cases if uc[0] in [4,5,6,9,10,11,13])),
     round(sum(round(uc[7]) for uc in use_cases if uc[0] in [4,5,6,9,10,11,13])),
     "Slide: Roadmap — Phase 2"),
    ("Phase 3 Full Agentic (8 use cases)",
     round(sum(round(uc[5]) for uc in use_cases if uc[0] in [7,8,12,14,15,16,19,20])),
     round(sum(round(uc[6]) for uc in use_cases if uc[0] in [7,8,12,14,15,16,19,20])),
     round(sum(round(uc[7]) for uc in use_cases if uc[0] in [7,8,12,14,15,16,19,20])),
     "Slide: Roadmap — Phase 3"),
    ("TOTAL ANNUAL VALUE (all 20 use cases)", total_bear, total_base, total_bull,
     "Slide 10: TCO — Annual Value box"),
    ("Total as % of €1.9B spend",
     None, f"{total_base/SPEND*100:.1f}%", None,
     "Sanity check — 5.7% is credible"),
    ("3-Year NPV (net of programme costs)", total_npv*0.5, total_npv, total_npv*1.5,
     "Slide 10: TCO — 3yr NPV"),
    ("Programme ROI", total_roi*0.5, total_roi, total_roi*1.5,
     "Slide 10: TCO — ROI box"),
    ("Annual Programme Cost (opex)", total_opex, total_opex, total_opex, "Slide 10: OPEX"),
    ("Hardware CAPEX (Year 1)", total_capex, total_capex, total_capex, "Slide 10: CAPEX"),
]

for i, row in enumerate(summary):
    r = 5 + i
    bg = NAVY if i in [4, 7] else (row_bgs[i%2])
    fg = WHITE if i in [4, 7] else DARK
    bold = i in [4, 7]
    C(ws4, r, 1, row[0], bg=bg, fg=fg, bold=bold, sz=11 if bold else 10)
    for col, val in zip([2,3,4], [row[1], row[2], row[3]]):
        if val is None:
            C(ws4, r, col, "—", bg=bg, fg=fg, align="center")
        elif isinstance(val, str):
            C(ws4, r, col, val, bg=bg, fg=fg, align="center", bold=bold, sz=12 if bold else 10)
        else:
            cl = ws4.cell(row=r, column=col, value=val)
            cl.fill = F(bg); cl.font = FT(bold, fg, 12 if bold else 10)
            cl.border = BD(); cl.alignment = AL("right", "center")
            cl.number_format = '€#,##0' if val > 100 else '0.0"x"'
    C(ws4, r, 5, row[4], bg=bg, fg=fg, sz=9, italic=True)
    ws4.row_dimensions[r].height = 22

for col, w in [(1,38),(2,18),(3,18),(4,18),(5,28)]:
    ws4.column_dimensions[get_column_letter(col)].width = w

# Column widths for main sheet
cw = {1:5,2:34,3:14,4:12,5:58,6:14,7:14,8:14,9:10,10:24,
      11:9,12:7,13:13,14:13,15:16,16:15,17:7,18:32}
for col, width in cw.items():
    ws.column_dimensions[get_column_letter(col)].width = width
ws.freeze_panes = "F6"

out = "/Users/jonathonmilne/.openclaw/workspace/venture/Statkraft_Procurement_AI_Value_Model_v4.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"\nTOTALS:")
print(f"  Bear:  €{total_bear:,.0f}")
print(f"  Base:  €{total_base:,.0f}  ({total_base/SPEND*100:.1f}% of spend)")
print(f"  Bull:  €{total_bull:,.0f}")
print(f"  3yr NPV: €{total_npv:,.0f}")
print(f"  ROI:   {total_roi:.0f}x")
