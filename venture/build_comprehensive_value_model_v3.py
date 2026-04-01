import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook()

# ── COLOURS ──────────────────────────────────────────────────────────────────
NAVY   = "002147"; STEEL  = "2B6CB0"; WHITE  = "FFFFFF"; GREY   = "718096"
AMBER  = "D97F06"; GREEN  = "27AE60"; RED    = "C0392B"; DARK   = "2D3A4A"
LTBLUE = "EBF4FF"; DEEP   = "041220"; PGREEN = "D4EDDA"; PAMBER = "FFF3CD"
PRED   = "F8D7DA"; TEAL   = "0B7A75"; YELLOW = "FFFACD"

def F(c): return PatternFill("solid", fgColor=c)
def FT(b=False, c=DARK, s=10, i=False): return Font(bold=b, color=c, size=s, name="Calibri", italic=i)
def AL(h="left", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)
def BD():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

def cell(ws, r, c, val, bg=WHITE, fg=DARK, bold=False, sz=10,
         align="left", wrap=False, italic=False, fmt=None):
    cl = ws.cell(row=r, column=c, value=val)
    cl.fill = F(bg); cl.font = FT(bold, fg, sz, italic)
    cl.alignment = AL(align, "center", wrap); cl.border = BD()
    if fmt: cl.number_format = fmt
    return cl

def hdr(ws, r, cols, bg=NAVY, fg=WHITE, sz=10):
    for i, (v, w) in enumerate(cols, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.fill = F(bg); c.font = FT(True, fg, sz)
        c.alignment = AL("center", "center", True); c.border = BD()
        if w: ws.column_dimensions[get_column_letter(i)].width = w

def sec(ws, r, title, ncols=20, bg=STEEL, fg=WHITE):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=title)
    c.fill = F(bg); c.font = FT(True, fg, 11)
    c.alignment = AL("left", "center"); c.border = BD()
    ws.row_dimensions[r].height = 20

def title_banner(ws, r1, r2, text, bg=NAVY, fg=WHITE, sz=14, ncols=20):
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=ncols)
    c = ws.cell(row=r1, column=1, value=text)
    c.fill = F(bg); c.font = FT(True, fg, sz)
    c.alignment = AL("center", "center"); c.border = BD()
    for r in range(r1, r2+1): ws.row_dimensions[r].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COMPREHENSIVE USE CASE VALUE REGISTER
# All use cases from Qwen 35B research + original 8 agents
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Use Case Value Register"
ws1.sheet_view.showGridLines = False

title_banner(ws1, 1, 2,
    "STATKRAFT PROCUREMENT AI — COMPREHENSIVE USE CASE VALUE REGISTER  |  Based on Qwen 35B Market Research + Internal Benchmarks",
    ncols=20)
ws1.merge_cells("A3:T3")
c = ws1.cell(row=3, column=1,
    value=f"Prepared: March 2026  |  Research Sources: McKinsey, BCG, Hackett Group, Equinor, Zip, Coupa, Resilinc, Ironclad, Fairmarkit, Arkestro, SAP, Jaggaer  |  All EUR values based on €500M annual spend, 45 FTE, €120K FTE cost")
c.fill = F(STEEL); c.font = FT(False, WHITE, 9, True); c.alignment = AL("center"); c.border = BD()

# ASSUMPTION INPUTS (yellow cells) ─────────────────────────────────────────
sec(ws1, 5, "  ⚙️  KEY INPUTS — Change yellow cells to reflect Statkraft actuals", ncols=20)
hdr(ws1, 6, [
    ("Input Parameter",32),("Value",12),("Unit",10),("Source",25),
    (None,0),(None,0),(None,0),(None,0),(None,0),(None,0),
    (None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0)
], bg=STEEL)

inputs = [
    ("Annual procurement spend (EUR)",          1_900_000_000, "EUR",   "Statkraft Annual Report 2024"),
    ("Procurement headcount (FTEs)",             45,          "FTEs",    "Statkraft internal"),
    ("Avg FTE fully-loaded cost (EUR/yr)",       120_000,     "EUR/yr",  "Statkraft HR"),
    ("Sourcing events per year",                 250,         "events",  "Statkraft estimate"),
    ("Avg hours per sourcing event (RFQ→award)", 40,          "hrs",     "Industry avg"),
    ("Active contracts in portfolio",            1200,        "contracts","Statkraft estimate"),
    ("Avg hours to manually review contract",    4,           "hrs",     "Luminance benchmark"),
    ("Suppliers actively monitored",             200,         "count",   "Statkraft target"),
    ("Annual manual supplier review cost",       150_000,     "EUR/yr",  "Statkraft estimate"),
    ("Current spend classification accuracy",    0.65,        "%",       "Current state"),
    ("MRO/capex emergency premium (% of spend)", 0.03,        "%",       "Industry avg"),
    ("Avg hours to build negotiation pack",      8,           "hrs",     "Statkraft estimate"),
    ("Category strategies produced/year",        3,           "count",   "Statkraft estimate"),
    ("Total capex project spend (EUR/yr)",        400_000_000, "EUR/yr", "Statkraft estimate (approx 20% of total spend)"),
    ("Avg project cost overrun % (energy sector)",0.40,       "%",       "ScienceDirect 2025 — 662 projects"),
    ("CSRD compliance manual checks/supplier",   12,          "checks/yr","Regulatory requirement"),
    ("% tail spend (<€10K) as % of total spend", 0.08,        "%",       "Industry avg 8-15%"),
    ("Avg invoice processing time (mins)",        15,          "mins",    "Industry avg"),
    ("Invoices processed per year",              18_000,      "count",   "Statkraft estimate"),
    ("Supplier onboarding time (days)",           45,          "days",    "Statkraft estimate"),
]
INROW = 7  # Row where inputs start
for i, (label, val, unit, src) in enumerate(inputs):
    r = INROW + i
    cell(ws1, r, 1, label, bg=LTBLUE, bold=True)
    c = ws1.cell(row=r, column=2, value=val)
    c.fill = F(YELLOW); c.font = FT(True, DARK, 11); c.border = BD(); c.alignment = AL("center")
    if unit == "%" and val < 1: c.number_format = "0.0%"
    elif isinstance(val, int) and val > 1000: c.number_format = '#,##0'
    cell(ws1, r, 3, unit, bg=LTBLUE, align="center")
    cell(ws1, r, 4, src, bg=LTBLUE, fg=GREY, italic=True)

# Named references for formula convenience (row numbers)
# B7=spend, B8=FTEs, B9=FTE_cost, B10=sourcing_events, B11=hrs_per_event
# B12=contracts, B13=hrs_per_contract, B14=suppliers, B15=supplier_review_cost
# B16=spend_accuracy, B17=MRO_pct, B18=negotiation_hrs, B19=strategies/yr
# B20=capex_spend, B21=overrun_pct, B22=CSRD_checks, B23=tail_pct, B24=invoice_mins
# B25=invoices, B26=onboarding_days

# USE CASE VALUE TABLE ─────────────────────────────────────────────────────────
SEC_ROW = INROW + len(inputs) + 2  # = 29
sec(ws1, SEC_ROW, "  💰  USE CASE VALUE ANALYSIS — All 20 Use Cases with Evidence-Based Benchmarks", ncols=20)

hdr(ws1, SEC_ROW+1, [
    ("#",4), ("Use Case",30), ("Category",14),
    ("Benchmark Source",22), ("Benchmark %/Metric",16),
    ("Calculation Method",35),
    ("Bear\n(50% of base)",14), ("Base\nEstimate",14), ("Bull\n(150% of base)",14),
    ("Confidence",10), ("Evidence Quality",14),
    ("Effort\n(weeks)",8), ("Time to\nValue",10),
    ("Track",6), ("One-time\nCost (EUR)",12), ("Annual\nCost (EUR)",12),
    ("Net Annual\nValue (EUR)",14), ("3yr NPV\n(EUR)",14),
    ("ROI",6), ("Notes / Assumptions",35)
], bg=NAVY, fg=WHITE, sz=9)

DR = SEC_ROW + 2  # data start row

# All 20 use cases — formulas reference B7:B26
use_cases = [
    # [id, name, cat, source, benchmark, calc_method,
    #  bear_formula, base_formula, bull_formula, confidence, evidence, effort, ttv, track, capex, opex, notes]
    (1,"Spend Classification & Enrichment","Data Quality",
     "McKinsey / Equinor (500M NOK automation)",
     "5-10% spend recovery from better classification accuracy",
     "Spend × (target accuracy 88% - current 65%) × 5% recovery rate",
     f"=B7*(0.88-B{INROW+9})*0.025", f"=B7*(0.88-B{INROW+9})*0.05", f"=B7*(0.88-B{INROW+9})*0.075",
     "High","McKinsey confirmed, Equinor implemented",3,"4–6 wks","B",0,15000,
     "Cleans data for all other AI use cases. Foundation use case. SAP Joule native."),

    (2,"Contract NLP Scanning & Review","Contract Mgmt",
     "Luminance / Icertis benchmark — 31% already using AI",
     "80% faster contract review; review 1,200 contracts/yr",
     "Hours saved × FTE cost ÷ 1750 working hrs per year",
     f"=B{INROW+2}/1750*B{INROW+6}*B{INROW+12}*0.8*0.4",
     f"=B{INROW+2}/1750*B{INROW+6}*B{INROW+12}*0.8",
     f"=B{INROW+2}/1750*B{INROW+6}*B{INROW+12}*0.8*1.5",
     "High","Luminance/Icertis benchmarked, 31% adoption live",4,"6–8 wks","B",5000,8000,
     "80% faster = 3.2 hrs saved per contract review. 1,200 contracts × 3.2h = 3,840h saved."),

    (3,"Supplier Monitoring Agent (Top 50→200)","Risk Mgmt",
     "Resilinc (Gartner Magic Quadrant Leader) / Shell",
     "70% reduction in manual monitoring costs; early warning reduces incidents",
     "Manual monitoring cost reduction + incident prevention value",
     f"=B{INROW+8}*0.7*0.5", f"=B{INROW+8}*0.7", f"=B{INROW+8}*0.7*1.5",
     "High","Shell/TotalEnergies live, Resilinc Gartner Leader",6,"8–10 wks","D",10000,25000,
     "Continuous vs. annual review. Resilinc: real-time disruption monitoring, tariff screening, forced labor compliance."),

    (4,"RFQ & Tender Drafting Agent","Sourcing",
     "Ørsted / BP — autonomous RFQ generation from spec",
     "75% time reduction; 4–8h saved per event",
     "Hours saved per event × events/yr × FTE cost ÷ 1750",
     f"=B{INROW+4}*B{INROW+3}*0.75*B{INROW+2}/1750*0.4",
     f"=B{INROW+4}*B{INROW+3}*0.75*B{INROW+2}/1750",
     f"=B{INROW+4}*B{INROW+3}*0.75*B{INROW+2}/1750*1.5",
     "High","Ørsted and BP live, Fairmarkit: 10x events per FTE",5,"8–12 wks","D",15000,30000,
     "250 events × 30h saved × €68.57/hr = €514K base. Fairmarkit: $40K savings per buyer per week."),

    (5,"Demand Forecasting — MRO & Capex","MRO/Supply",
     "Vattenfall (6–12m horizon) / McKinsey (35% inventory reduction)",
     "50% reduction in emergency premium buys; 13% inventory reduction",
     "Emergency premium spend reduction + inventory carrying cost savings",
     f"=B7*B{INROW+10}*0.5*0.4+B7*0.02*0.13*0.4",
     f"=B7*B{INROW+10}*0.5+B7*0.02*0.13",
     f"=B7*B{INROW+10}*0.5*1.5+B7*0.02*0.13*1.5",
     "High","Vattenfall live, McKinsey 35% inventory reduction confirmed",8,"12–16 wks","D",20000,40000,
     "MRO emergency premium 3% of spend = €15M. 50% reduction = €7.5M. Inventory saving ~€1.3M additional."),

    (6,"Negotiation Intelligence & Prep","Sourcing",
     "Arkestro (18.8% savings per $1M) / McKinsey should-cost 13% savings",
     "0.3% additional savings from better-prepared negotiations",
     "Total spend × % additional savings from AI-assisted negotiation prep",
     f"=B7*0.003*0.5", f"=B7*0.003", f"=B7*0.003*1.8",
     "Medium","Arkestro: 18.8% savings/event; McKinsey: 13% should-cost savings",5,"8–12 wks","D",10000,20000,
     "Conservative: 0.3% of €500M = €1.5M. Arkestro achieving 18.8% per event on direct materials negotiated."),

    (7,"Category Strategy Agent","Strategy",
     "McKinsey autonomous category agents 15–30% efficiency; BCG 15–45% cost reduction",
     "0.5% spend savings from better category strategy coverage (3→12 strategies/yr)",
     "Spend × savings % enabled by broader category strategy coverage",
     f"=B7*0.005*0.4", f"=B7*0.005", f"=B7*0.005*2",
     "Medium","McKinsey/BCG benchmarked; Statkraft currently 3 strategies/yr",8,"12–20 wks","D",25000,50000,
     "Increase from 3 to 12 category strategies per year. Each strategy covers average 4% of spend. 0.5% savings rate."),

    (8,"CSRD & ESG Compliance Automation","Compliance",
     "CSRD mandatory 2026; Resilinc forced labor / tariff screening live",
     "80% reduction in manual compliance checking hours",
     "Manual compliance hours saved × FTE cost ÷ 1750 + penalty risk reduction",
     f"=B{INROW+2}/1750*B{INROW+1}*B{INROW+15}*0.8*0.4",
     f"=B{INROW+2}/1750*B{INROW+1}*B{INROW+15}*0.8",
     f"=B{INROW+2}/1750*B{INROW+1}*B{INROW+15}*0.8*1.5",
     "High","CSRD mandatory, Resilinc agent live on Azure",5,"10–14 wks","D",15000,25000,
     "200 suppliers × 12 checks/yr × 0.8 hours avg = 1,920h saved. CSRD audit penalty risk reduction = additional strategic value."),

    (9,"Invoice Processing & Anomaly Detection","P2P",
     "Jaggaer (anomaly detection live) / SAP Joule (Q4 2025)",
     "70% reduction in manual invoice processing time",
     "Invoices × time saved × FTE cost rate",
     f"=B{INROW+18}*B{INROW+17}/60*B{INROW+2}/1750*0.7*0.4",
     f"=B{INROW+18}*B{INROW+17}/60*B{INROW+2}/1750*0.7",
     f"=B{INROW+18}*B{INROW+17}/60*B{INROW+2}/1750*0.7*1.5",
     "High","Jaggaer anomaly detection launched 2025; SAP Joule native",3,"4–6 wks","B",0,5000,
     "18,000 invoices × 15min × 70% saving × €68.57/hr = €121K. Plus fraud/error prevention value."),

    (10,"Market Intelligence Agent (Daily Briefings)","Market Intel",
     "GEP (15% lower logistics costs) / Equinor commodity monitoring",
     "1% better purchase timing on commodity-exposed spend (est. 30% of total)",
     "Commodity-exposed spend × price timing improvement",
     f"=B7*0.3*0.01*0.4", f"=B7*0.3*0.01", f"=B7*0.3*0.01*2",
     "Medium","GEP/McKinsey benchmarked; Equinor live on commodity monitoring",2,"2–4 wks","A/B",0,3000,
     "30% of spend commodity-exposed = €150M. 1% better timing = €1.5M base. Conservative given steel/copper volatility 2024-25."),

    (11,"Should-Cost Modelling Agent","Cost Intelligence",
     "McKinsey: 13% savings in raw-materials from should-cost AI",
     "1% savings on addressable direct spend via better cost baselines",
     "Direct material spend × should-cost savings rate",
     f"=B7*0.25*0.01*0.4", f"=B7*0.25*0.01", f"=B7*0.25*0.013",
     "High","McKinsey: 13% raw materials savings confirmed; Arkestro: 18.8%",6,"8–12 wks","D",20000,35000,
     "25% of spend = direct materials. 1% conservative savings (McKinsey 13%, using 1% for Statkraft first year)."),

    (12,"Supplier Onboarding Automation","Supplier Mgmt",
     "TealBook / Scoutbee (Coupa acquisition Oct 2025)",
     "60% reduction in onboarding time (45 days → 18 days)",
     "Time saved × FTE cost + faster project start value",
     f"=B{INROW+2}/250*B{INROW+19}*0.6*40*0.4",
     f"=B{INROW+2}/250*B{INROW+19}*0.6*40",
     f"=B{INROW+2}/250*B{INROW+19}*0.6*40*1.5",
     "Medium","TealBook/Scoutbee live; 60% onboarding time reduction typical",6,"10–14 wks","C",10000,15000,
     "Assume 40 new suppliers/yr, 45 days → 18 days saved × €460/day FTE cost."),

    (13,"Tail Spend Automation Agent","P2P",
     "Fairmarkit: $40K savings per buyer/week; 10x events per FTE",
     "15% savings on tail spend (typically managed at sub-optimal prices)",
     "Tail spend × savings rate from competitive mini-tendering",
     f"=B7*B{INROW+16}*0.15*0.4", f"=B7*B{INROW+16}*0.15", f"=B7*B{INROW+16}*0.15*1.8",
     "High","Fairmarkit live with documented $40K/buyer/week savings",4,"6–10 wks","D",10000,20000,
     "Tail spend = 8% of €500M = €40M. 15% savings via competitive mini-tenders = €6M. Conservative vs Fairmarkit benchmarks."),

    (14,"Project Cost Intelligence Agent","Decision Intel",
     "ScienceDirect (40% overrun avg) / McKinsey (80% on megaprojects)",
     "2% improvement in estimate accuracy on capex spend",
     "Capex × overrun rate × % reduction from better cost intelligence",
     f"=B{INROW+13}*B{INROW+14}*0.02*0.3",
     f"=B{INROW+13}*B{INROW+14}*0.02",
     f"=B{INROW+13}*B{INROW+14}*0.02*2",
     "Medium","ScienceDirect: 40% avg overrun (662 projects, $1.358T). Conservative 2% improvement.",10,"16–24 wks","D",30000,60000,
     "€200M capex × 40% overrun rate × 2% improvement = €1.6M base. Mines SAP/Jaggaer actuals vs. estimates."),

    (15,"EPC Contract AI Review Agent","Contract Mgmt",
     "Ironclad ($150M ARR, 39% growth) / Icertis (33 Fortune 100 clients)",
     "40–60% legal cycle time reduction on complex contracts",
     "Legal/procurement hours saved on EPC contract review per year",
     f"=B{INROW+2}/1750*20*B{INROW+5}*40*0.4*0.4",
     f"=B{INROW+2}/1750*20*B{INROW+5}*40*0.4",
     f"=B{INROW+2}/1750*20*B{INROW+5}*40*0.6",
     "Medium","Ironclad/Icertis live at Fortune 100; 40% cycle time reduction confirmed",8,"12–18 wks","D",20000,40000,
     "20 major EPC contracts/yr × avg 40 review hours × 40% savings. Plus risk reduction from catching non-standard clauses."),

    (16,"Supplier Risk Intelligence — Financial","Risk Mgmt",
     "Resilinc (Gartner Magic Quadrant Leader) / Shell continuous monitoring",
     "1 major supplier failure prevented per 2 years (€2M avg cost prevented)",
     "Probability-weighted cost of supplier failure events avoided",
     f"=2000000/2*0.3", f"=2000000/2", f"=2000000/2*2",
     "Medium","Resilinc Gartner Leader; Shell financial monitoring live",6,"8–12 wks","D",10000,20000,
     "€2M avg cost of unplanned supplier switch (emergency sourcing + project delay). 1 event prevented per 2 yrs conservative."),

    (17,"SAP Joule Activation (Native — Quick Win)","Platform AI",
     "SAP: 30% productivity target; 14 new Joule agents Q4 2025",
     "20% productivity gain on SAP-native procurement tasks",
     "SAP users × time in SAP × productivity gain",
     f"=B{INROW+2}/1750*1000*0.2*B{INROW+2}*0.4",
     f"=B{INROW+2}/1750*1000*0.2*B{INROW+2}",
     f"=B{INROW+2}/1750*1000*0.2*B{INROW+2}*1.5",
     "High","SAP native — no custom integration. Joule agents live Q4 2025",1,"1–2 wks","A",0,0,
     "Included in existing SAP licence. Bid analysis, SOW automation, NL demand routing, delivery date mass updates."),

    (18,"Jaggaer JAI Activation (Native)","Platform AI",
     "Jaggaer JAI: Gartner score 3.75; copilot + anomaly detection live",
     "15% efficiency gain on sourcing and contract tasks in Jaggaer",
     "Jaggaer users × time × efficiency gain",
     f"=B{INROW+2}/1750*800*0.15*B{INROW+2}*0.4",
     f"=B{INROW+2}/1750*800*0.15*B{INROW+2}",
     f"=B{INROW+2}/1750*800*0.15*B{INROW+2}*1.5",
     "High","Native Jaggaer JAI — copilot live, agentic platform 2026",1,"1–2 wks","A",0,0,
     "Included in existing Jaggaer licence. JAI copilot, anomaly detection, conversational CLM live."),

    (19,"Autonomous Tender Management (Low Value)","P2P",
     "Zip (50+ agents, $6B savings) / Pactum (<$150K autonomous negotiation)",
     "End-to-end automation of tenders <€50K",
     "Low-value tender volume × avg value × savings rate",
     f"=B{INROW+3}*0.4*50000*0.08*0.4",
     f"=B{INROW+3}*0.4*50000*0.08",
     f"=B{INROW+3}*0.4*50000*0.08*1.5",
     "Medium","Zip and Pactum live at scale; 100 events automated per year est.",6,"10–16 wks","D",15000,30000,
     "40% of 250 events = 100 low-value. Avg €50K × 8% savings = €400K base. Pactum managing <$150K autonomously."),

    (20,"Supply Chain Network Risk Mapping","Risk/Strategy",
     "GEP (15% lower logistics, 35% inventory) / Resilinc sub-tier mapping",
     "Risk-adjusted value: 1 major supply disruption prevented per year",
     "Probability × impact of supply disruption avoided",
     f"=5000000*0.15*0.3", f"=5000000*0.15", f"=5000000*0.15*2",
     "Low","GEP/Resilinc benchmarks; supply disruption costs well documented",10,"16–24 wks","D",25000,50000,
     "Major supply disruption (e.g. transformer shortage) costs €5M+ in project delay. 15% probability prevention."),
]

row_cols = [LTBLUE, WHITE]
for i, uc in enumerate(use_cases):
    r = DR + i
    bg = row_cols[i % 2]
    (uc_id, name, cat, source, bench, calc,
     bear, base, bull, conf, evid, effort, ttv, track, capex, opex, notes) = uc

    cell(ws1, r, 1, uc_id, bg=STEEL, fg=WHITE, bold=True, align="center", sz=11)
    cell(ws1, r, 2, name, bg=bg, bold=True, sz=10)
    cell(ws1, r, 3, cat, bg=bg, align="center", sz=9)
    cell(ws1, r, 4, source, bg=bg, fg=GREY, sz=8, italic=True, wrap=True)
    cell(ws1, r, 5, bench, bg=bg, sz=9, wrap=True)
    cell(ws1, r, 6, calc, bg=bg, fg=GREY, sz=8, italic=True, wrap=True)

    # Bear / Base / Bull — live formulas
    for col_idx, formula in [(7, bear), (8, base), (9, bull)]:
        c = ws1.cell(row=r, column=col_idx, value=formula)
        colors = {7: PRED, 8: PGREEN, 9: "D1F2EB"}
        c.fill = F(colors[col_idx])
        c.font = FT(True, DARK if col_idx==7 else GREEN, 10)
        c.border = BD(); c.alignment = AL("center")
        c.number_format = '€#,##0'

    # Confidence
    conf_bg = {"High": PGREEN, "Medium": PAMBER, "Low": PRED}.get(conf, WHITE)
    conf_fg = {"High": GREEN, "Medium": AMBER, "Low": RED}.get(conf, DARK)
    cell(ws1, r, 10, conf, bg=conf_bg, fg=conf_fg, bold=True, align="center", sz=9)
    cell(ws1, r, 11, evid, bg=bg, sz=8, italic=True, wrap=True)
    cell(ws1, r, 12, effort, bg=bg, align="center", sz=9)
    cell(ws1, r, 13, ttv, bg=bg, align="center", sz=9)
    cell(ws1, r, 14, track, bg=NAVY if track=="D" else (STEEL if track=="C" else LTBLUE),
         fg=WHITE if track in ["D","C"] else NAVY, bold=True, align="center", sz=10)

    # Costs
    for col_idx, val, fmt in [(15, capex, '€#,##0'), (16, opex, '€#,##0/yr')]:
        c = ws1.cell(row=r, column=col_idx, value=val)
        c.fill = F(PAMBER); c.font = FT(False, AMBER); c.border = BD()
        c.alignment = AL("center"); c.number_format = fmt

    # Net Annual Value = Base - Annual Cost
    net_formula = f"=H{r}-P{r}"
    c = ws1.cell(row=r, column=17, value=net_formula)
    c.fill = F(PGREEN); c.font = FT(True, GREEN, 11); c.border = BD()
    c.alignment = AL("center"); c.number_format = '€#,##0'

    # 3yr NPV = Net × 3 - CAPEX
    npv_formula = f"=Q{r}*3-O{r}"
    c = ws1.cell(row=r, column=18, value=npv_formula)
    c.fill = F(PGREEN); c.font = FT(True, GREEN, 11); c.border = BD()
    c.alignment = AL("center"); c.number_format = '€#,##0'

    # ROI = Base / (Capex + Opex) if costs > 0
    roi_formula = f"=IF(O{r}+P{r}>0,H{r}/(O{r}+P{r}),\"∞\")"
    c = ws1.cell(row=r, column=19, value=roi_formula)
    c.fill = F(PGREEN); c.font = FT(True, GREEN, 11); c.border = BD()
    c.alignment = AL("center"); c.number_format = '0.0"x"'

    cell(ws1, r, 20, notes, bg=bg, fg=GREY, sz=8, italic=True, wrap=True)
    ws1.row_dimensions[r].height = 45

# TOTALS ROW ─────────────────────────────────────────────────────────────────
TOTAL_ROW = DR + len(use_cases)
ws1.merge_cells(f"A{TOTAL_ROW}:F{TOTAL_ROW}")
c = ws1.cell(row=TOTAL_ROW, column=1, value="PROGRAMME TOTAL (all 20 use cases fully deployed)")
c.fill = F(NAVY); c.font = FT(True, WHITE, 12); c.border = BD(); c.alignment = AL("right")

for col_idx, col_letter in [(7,"G"), (8,"H"), (9,"I")]:
    c = ws1.cell(row=TOTAL_ROW, column=col_idx,
                 value=f"=SUM({col_letter}{DR}:{col_letter}{TOTAL_ROW-1})")
    c.fill = F(NAVY); c.font = FT(True, WHITE, 13); c.border = BD()
    c.alignment = AL("center"); c.number_format = '€#,##0'

for col_idx, col_letter in [(15,"O"), (16,"P"), (17,"Q"), (18,"R")]:
    c = ws1.cell(row=TOTAL_ROW, column=col_idx,
                 value=f"=SUM({col_letter}{DR}:{col_letter}{TOTAL_ROW-1})")
    c.fill = F(NAVY); c.font = FT(True, WHITE, 13); c.border = BD()
    c.alignment = AL("center"); c.number_format = '€#,##0'

# Programme ROI
roi_c = ws1.cell(row=TOTAL_ROW, column=19,
                  value=f"=H{TOTAL_ROW}/(O{TOTAL_ROW}+P{TOTAL_ROW})")
roi_c.fill = F(NAVY); roi_c.font = FT(True, WHITE, 13); roi_c.border = BD()
roi_c.alignment = AL("center"); roi_c.number_format = '0.0"x"'

ws1.row_dimensions[TOTAL_ROW].height = 28

# Phased rollout summary ──────────────────────────────────────────────────────
PHASE_ROW = TOTAL_ROW + 2
sec(ws1, PHASE_ROW, "  📅  PHASED ROLLOUT SUMMARY — Expected value by phase", ncols=20)
hdr(ws1, PHASE_ROW+1, [
    ("Phase",14),("Use Cases",30),("Timeline",10),
    ("Bear Value",14),("Base Value",14),("Bull Value",14),
    ("Cumulative Base",16),(None,0),(None,0),(None,0),
    (None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0),(None,0)
], bg=STEEL)

phases = [
    ("Phase 0 — Quick Wins","#17 SAP Joule, #18 Jaggaer JAI","Weeks 1–2",
     f"=H{DR+16}+H{DR+17}", f"=H{DR+16}+H{DR+17}", f"=H{DR+16}+H{DR+17}"),
    ("Phase 1 — First Pilots","#1 Spend Classify, #2 Contract NLP, #3 Supplier Monitor","Weeks 3–12",
     f"=G{DR}+G{DR+1}+G{DR+2}", f"=H{DR}+H{DR+1}+H{DR+2}", f"=I{DR}+I{DR+1}+I{DR+2}"),
    ("Phase 2 — Expand","#4 RFQ, #5 Demand Forecast, #6 Negotiation, #9 Invoice, #10 Market Intel, #13 Tail Spend","Months 4–6",
     f"=G{DR+3}+G{DR+4}+G{DR+5}+G{DR+8}+G{DR+9}+G{DR+12}",
     f"=H{DR+3}+H{DR+4}+H{DR+5}+H{DR+8}+H{DR+9}+H{DR+12}",
     f"=I{DR+3}+I{DR+4}+I{DR+5}+I{DR+8}+I{DR+9}+I{DR+12}"),
    ("Phase 3 — Full Agentic","#7 Category Strategy, #8 CSRD, #11 Should-Cost, #14 Decision Intel, #15 EPC Review, #16 Risk, #19 Auto-Tender, #20 Network","Months 7–12",
     f"=G{DR+6}+G{DR+7}+G{DR+10}+G{DR+13}+G{DR+14}+G{DR+15}+G{DR+18}+G{DR+19}",
     f"=H{DR+6}+H{DR+7}+H{DR+10}+H{DR+13}+H{DR+14}+H{DR+15}+H{DR+18}+H{DR+19}",
     f"=I{DR+6}+I{DR+7}+I{DR+10}+I{DR+13}+I{DR+14}+I{DR+15}+I{DR+18}+I{DR+19}"),
    ("FULL PROGRAMME","All 20 use cases deployed","Year 1+",
     f"=G{TOTAL_ROW}", f"=H{TOTAL_ROW}", f"=I{TOTAL_ROW}"),
]
phase_bgs = [LTBLUE, WHITE, LTBLUE, WHITE, NAVY]
phase_fgs = [DARK, DARK, DARK, DARK, WHITE]
for i, (phase, uc, timeline, bear_f, base_f, bull_f) in enumerate(phases):
    r = PHASE_ROW + 2 + i
    bg = phase_bgs[i]; fg = phase_fgs[i]
    bold = (i == 4)
    cell(ws1, r, 1, phase, bg=bg, fg=fg, bold=bold, sz=10 if not bold else 12)
    cell(ws1, r, 2, uc, bg=bg, fg=fg, sz=9)
    cell(ws1, r, 3, timeline, bg=bg, fg=fg, align="center", sz=9)
    for col_idx, formula in [(4, bear_f), (5, base_f), (6, bull_f)]:
        c = ws1.cell(row=r, column=col_idx, value=formula)
        c.fill = F(bg if not bold else NAVY)
        c.font = FT(bold, fg if not bold else WHITE, 11 if bold else 10)
        c.border = BD(); c.alignment = AL("center"); c.number_format = '€#,##0'
    ws1.row_dimensions[r].height = 22

# Cumulative column
CUM_ROW_START = PHASE_ROW + 2
for j in range(len(phases)):
    r = CUM_ROW_START + j
    if j == 0:
        c = ws1.cell(row=r, column=7, value=f"=E{r}")
    else:
        c = ws1.cell(row=r, column=7, value=f"=G{r-1}+E{r}")
    c.fill = F(PGREEN); c.font = FT(True, GREEN, 11); c.border = BD()
    c.alignment = AL("center"); c.number_format = '€#,##0'

# Column widths
col_widths = {1:5,2:32,3:14,4:24,5:18,6:38,7:15,8:15,9:15,
              10:10,11:16,12:9,13:11,14:7,15:13,16:13,17:16,18:15,19:7,20:38}
for col, width in col_widths.items():
    ws1.column_dimensions[get_column_letter(col)].width = width

ws1.freeze_panes = f"C{DR}"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PROGRAMME SUMMARY (for PPTX update)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🎯 Summary for Deck")
ws2.sheet_view.showGridLines = False
title_banner(ws2, 1, 2, "PROGRAMME VALUE SUMMARY — Numbers to Update in Presentation", ncols=8)

hdr(ws2, 4, [("Metric",35),("Bear",14),("Base",14),("Bull",14),
              ("Use in Deck",20),(None,0),(None,0),(None,0)], bg=NAVY)

PR = PHASE_ROW  # phase row reference in ws1
summary_rows = [
    ("Phase 0 Quick Wins (SAP Joule + Jaggaer JAI)",
     f"='📊 Use Case Value Register'!D{CUM_ROW_START}",
     f"='📊 Use Case Value Register'!E{CUM_ROW_START}",
     f"='📊 Use Case Value Register'!F{CUM_ROW_START}",
     "Year 1 fast start"),
    ("Phase 1 Pilots (Spend, Contract, Supplier Monitor)",
     f"='📊 Use Case Value Register'!D{CUM_ROW_START+1}",
     f"='📊 Use Case Value Register'!E{CUM_ROW_START+1}",
     f"='📊 Use Case Value Register'!F{CUM_ROW_START+1}",
     "90-day review gate"),
    ("Phase 2 Expanded (6 use cases)",
     f"='📊 Use Case Value Register'!D{CUM_ROW_START+2}",
     f"='📊 Use Case Value Register'!E{CUM_ROW_START+2}",
     f"='📊 Use Case Value Register'!F{CUM_ROW_START+2}",
     "Month 4-6 value"),
    ("Phase 3 Full Agentic (8 use cases)",
     f"='📊 Use Case Value Register'!D{CUM_ROW_START+3}",
     f"='📊 Use Case Value Register'!E{CUM_ROW_START+3}",
     f"='📊 Use Case Value Register'!F{CUM_ROW_START+3}",
     "Month 7-12 value"),
    ("TOTAL ANNUAL VALUE — ALL 20 USE CASES",
     f"='📊 Use Case Value Register'!G{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!H{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!I{TOTAL_ROW}",
     "Update Slide 10 TCO box"),
    ("Total Programme CAPEX",
     f"='📊 Use Case Value Register'!O{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!O{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!O{TOTAL_ROW}",
     "Update hardware cost slide"),
    ("Total Annual OPEX",
     f"='📊 Use Case Value Register'!P{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!P{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!P{TOTAL_ROW}",
     "Update TCO slide"),
    ("3-Year NPV",
     f"='📊 Use Case Value Register'!R{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!R{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!R{TOTAL_ROW}",
     "Update ROI box"),
    ("Programme ROI",
     f"='📊 Use Case Value Register'!S{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!S{TOTAL_ROW}",
     f"='📊 Use Case Value Register'!S{TOTAL_ROW}",
     "Update ROI box"),
]

row_bgs2 = [LTBLUE, WHITE]
for i, (metric, bear, base, bull, note) in enumerate(summary_rows):
    r = 5 + i
    bg = NAVY if i >= 4 else row_bgs2[i%2]
    fg = WHITE if i >= 4 else DARK
    bold = (i >= 4)
    cell(ws2, r, 1, metric, bg=bg, fg=fg, bold=bold, sz=11 if bold else 10)
    for col_idx, formula in [(2,bear),(3,base),(4,bull)]:
        c = ws2.cell(row=r, column=col_idx, value=formula)
        c.fill = F(bg if i<4 else NAVY)
        c.font = FT(bold, fg if i<4 else WHITE, 12 if bold else 10)
        c.border = BD(); c.alignment = AL("center"); c.number_format = '€#,##0'
    cell(ws2, r, 5, note, bg=bg, fg=fg, italic=True, sz=9)
    ws2.row_dimensions[r].height = 22 if bold else 18

for col, width in [(1,38),(2,16),(3,16),(4,16),(5,22)]:
    ws2.column_dimensions[get_column_letter(col)].width = width

out = "/Users/jonathonmilne/.openclaw/workspace/venture/Statkraft_Procurement_AI_Value_Model_v3.xlsx"
wb.save(out)
print(f"Excel saved: {out}")

# Print summary numbers for deck update
print("\n=== NUMBERS TO UPDATE IN DECK (Base estimates) ===")
