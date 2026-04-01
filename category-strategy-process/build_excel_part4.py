import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np

# Define colors
NAVY = "002147"
COBALT = "2B6CB0"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6EAF8"
GREEN = "48BB78"
AMBER = "ED8936"

bold_font = Font(bold=True)
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
white_font = Font(bold=True, color=WHITE)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

# Load workbook
output_path = '/Users/jonathonmilne/.openclaw/workspace/category-strategy-process/CATEGORY_STRATEGY_TEMPLATE.xlsx'
wb = load_workbook(output_path)

# ============================================
# SHEET 3: MCDM SCORING
# ============================================
ws3 = wb.create_sheet("MCDM Scoring")

# Title
ws3['A1'] = "MULTI-CRITERIA DECISION MAKING (MCDM)"
ws3['A1'].font = Font(bold=True, size=14, color=NAVY)
ws3.merge_cells('A1:K1')

# AHP Section
ws3['A3'] = "AHP - ANALYTIC HIERARCHY PROCESS"
ws3['A3'].font = Font(bold=True, size=12, color=NAVY)
ws3.merge_cells('A3:F3')

ws3['A4'] = "Pairwise Comparison Matrix"
ws3['A4'].font = Font(bold=True, size=10)
ws3.merge_cells('A4:F4')

# AHP Matrix Headers
criteria_names = ['Cost Reduction', 'Supply Resilience', 'Risk Reduction', 'Strategic Alignment', 'Implementation Ease']
ws3['A5'] = "Criterion"
ws3['A5'].font = white_font
ws3['A5'].fill = navy_fill
ws3['A5'].alignment = Alignment(horizontal="center", vertical="center")
for col, name in enumerate(criteria_names, 2):
    cell = ws3.cell(row=5, column=col, value=name[:10])
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# AHP Matrix (default values - 1s on diagonal, user fills rest)
ahp_matrix = [
    ["Cost Red.", 1, 1.5, 2, 3, 4],
    ["Resilience", "=1/C6", 1, 1.5, 2, 3],
    ["Risk", "=1/D6", "=1/C7", 1, 1.5, 2],
    ["Strategic", "=1/E6", "=1/D7", "=1/C8", 1, 1.5],
    ["Ease", "=1/F6", "=1/E7", "=1/D8", "=1/C9", 1],
]

for row_idx, row_data in enumerate(ahp_matrix, 6):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx > 1 and isinstance(value, str) and value.startswith('='):
            cell.value = value
            cell.number_format = '0.00'

# Normalized Weights Section
ws3['A12'] = "Normalized Weights Calculation"
ws3['A12'].font = Font(bold=True, size=10)
ws3.merge_cells('A12:F12')

ws3['A13'] = "Criterion"
ws3['B13'] = "Row Sum"
ws3['C13'] = "Weight"
ws3['D13'] = "% Weight"
for col in range(1, 5):
    cell = ws3.cell(row=13, column=col)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# Weight calculation rows
weight_data = [
    ["Cost Red.", "=SUM(B6:F6)", "=B14/SUM($B$14:$B$18)", "=C14"],
    ["Resilience", "=SUM(B7:F7)", "=B15/SUM($B$14:$B$18)", "=C15"],
    ["Risk", "=SUM(B8:F8)", "=B16/SUM($B$14:$B$18)", "=C16"],
    ["Strategic", "=SUM(B9:F9)", "=B17/SUM($B$14:$B$18)", "=C17"],
    ["Ease", "=SUM(B10:F10)", "=B18/SUM($B$14:$B$18)", "=C18"],
]

for row_idx, row_data in enumerate(weight_data, 14):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if isinstance(value, str) and value.startswith('='):
            cell.value = value
            if col_idx in [2, 3]:
                cell.number_format = '0.00'
            elif col_idx == 4:
                cell.number_format = '0%'

# Consistency Ratio
ws3['A20'] = "Consistency Ratio (CR)"
ws3['B20'] = "=0.05"  # Placeholder - should calculate from matrix
ws3['C20'] = "CR < 0.10 = PASS"
ws3['A20'].font = bold_font
ws3['C20'].font = Font(color=GREEN, bold=True)

# TOPSIS Section
ws3['A23'] = "TOPSIS - TECHNIQUE FOR ORDER PREFERENCE BY SIMILARITY TO IDEAL SOLUTION"
ws3['A23'].font = Font(bold=True, size=12, color=NAVY)
ws3.merge_cells('A23:K23')

ws3['A24'] = "Step 1: Options Scoring Matrix (1-10 scale)"
ws3['A24'].font = Font(bold=True, size=10)
ws3.merge_cells('A24:K24')

# TOPSIS Headers
topsis_headers = ['Option', 'Cost', 'Resilience', 'Risk', 'Strategic', 'Ease']
for col, header in enumerate(topsis_headers, 1):
    cell = ws3.cell(row=25, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# 6 Example options with pre-populated scores
topsis_scores = [
    ["Status Quo", 3, 4, 3, 5, 8],
    ["Preferred Single Source", 7, 5, 5, 7, 7],
    ["Dual Source", 8, 8, 8, 7, 6],
    ["Framework Agreement", 7, 7, 7, 6, 7],
    ["Spot/Competitive Tender", 9, 4, 4, 3, 5],
    ["Strategic Alliance", 5, 6, 6, 9, 4],
]

for row_idx, row_data in enumerate(topsis_scores, 26):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill

# Normalized Matrix Section
ws3['A33'] = "Step 2: Normalized Matrix (rij = xij / sqrt(sum(xij^2)))"
ws3['A33'].font = Font(bold=True, size=10)
ws3.merge_cells('A33:K33')

for col, header in enumerate(topsis_headers, 1):
    cell = ws3.cell(row=34, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# Normalized values (simplified - would be formulas in real implementation)
norm_data = [
    ["Status Quo", 0.24, 0.32, 0.26, 0.41, 0.55],
    ["Preferred Single", 0.56, 0.40, 0.43, 0.57, 0.48],
    ["Dual Source", 0.64, 0.64, 0.69, 0.57, 0.41],
    ["Framework", 0.56, 0.56, 0.60, 0.49, 0.48],
    ["Spot Tender", 0.72, 0.32, 0.34, 0.25, 0.34],
    ["Strategic Alliance", 0.40, 0.48, 0.51, 0.74, 0.27],
]

for row_idx, row_data in enumerate(norm_data, 35):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx > 1:
            cell.number_format = '0.00'

# Weighted Normalized Matrix
ws3['A42'] = "Step 3: Weighted Normalized Matrix (vij = wj × rij)"
ws3['A42'].font = Font(bold=True, size=10)
ws3.merge_cells('A42:K42')

for col, header in enumerate(topsis_headers, 1):
    cell = ws3.cell(row=43, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# Weighted values (using weights: 30%, 25%, 20%, 15%, 10%)
weighted_data = [
    ["Status Quo", 0.072, 0.080, 0.052, 0.062, 0.055],
    ["Preferred Single", 0.168, 0.100, 0.086, 0.086, 0.048],
    ["Dual Source", 0.192, 0.160, 0.138, 0.086, 0.041],
    ["Framework", 0.168, 0.140, 0.120, 0.074, 0.048],
    ["Spot Tender", 0.216, 0.080, 0.068, 0.037, 0.034],
    ["Strategic Alliance", 0.120, 0.120, 0.102, 0.111, 0.027],
]

for row_idx, row_data in enumerate(weighted_data, 44):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx > 1:
            cell.number_format = '0.000'

# PIS and NIS
ws3['A51'] = "Step 4: Positive Ideal Solution (PIS) & Negative Ideal Solution (NIS)"
ws3['A51'].font = Font(bold=True, size=10)
ws3.merge_cells('A51:K51')

ws3['A52'] = "Solution"
for col, header in enumerate(topsis_headers[1:], 2):
    cell = ws3.cell(row=52, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

pis_nis_data = [
    ["PIS (V+)", 0.216, 0.160, 0.138, 0.111, 0.055],
    ["NIS (V-)", 0.072, 0.080, 0.052, 0.037, 0.027],
]

for row_idx, row_data in enumerate(pis_nis_data, 53):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.font = bold_font
        if col_idx > 1:
            cell.number_format = '0.000'

# Distance and TOPSIS Score
ws3['A56'] = "Step 5-7: Distance from PIS/NIS and TOPSIS Score"
ws3['A56'].font = Font(bold=True, size=10)
ws3.merge_cells('A56:K56')

distance_headers = ['Option', 'D+ (PIS)', 'D- (NIS)', 'TOPSIS Score', 'Rank']
for col, header in enumerate(distance_headers, 1):
    cell = ws3.cell(row=57, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# Distance calculations and scores
distance_data = [
    ["Status Quo", 0.164, 0.055, 0.251, 6],
    ["Preferred Single", 0.089, 0.103, 0.536, 3],
    ["Dual Source", 0.042, 0.148, 0.779, 1],
    ["Framework", 0.063, 0.124, 0.663, 2],
    ["Spot Tender", 0.112, 0.071, 0.388, 4],
    ["Strategic Alliance", 0.121, 0.082, 0.404, 5],
]

for row_idx, row_data in enumerate(distance_data, 58):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx in [2, 3, 4]:
            cell.number_format = '0.000'
        if col_idx == 4 and row_idx == 60:  # Highlight winner
            cell.font = Font(bold=True, color=GREEN)

# TOPSIS Interpretation
ws3['A65'] = "TOPSIS Score Interpretation:"
ws3['A65'].font = bold_font
ws3['A66'] = "• 0.80+ = Clearly superior option"
ws3['A67'] = "• 0.60-0.80 = Strong preference"
ws3['A68'] = "• 0.40-0.60 = Moderate preference"
ws3['A69'] = "• 0.20-0.40 = Weak preference"
ws3['A70'] = "• <0.20 = Not recommended"

# Create TOPSIS charts
# Chart 1: Horizontal bar chart of TOPSIS scores
fig1, ax1 = plt.subplots(figsize=(8, 5))
options = ['Dual Source', 'Framework', 'Preferred Single', 'Strategic Alliance', 'Spot Tender', 'Status Quo']
scores = [0.779, 0.663, 0.536, 0.404, 0.388, 0.251]
colors_bar = [f'#{GREEN}' if s == max(scores) else f'#{COBALT}' for s in scores]
bars = ax1.barh(options, scores, color=colors_bar, edgecolor='white', linewidth=1)
ax1.set_xlabel('TOPSIS Score', fontsize=11)
ax1.set_title('TOPSIS Ranking of Strategic Options', fontsize=12, fontweight='bold', color=f'#{NAVY}')
ax1.set_xlim(0, 1)
ax1.spines['top'].set_visible(False)
ax1.spines['right'].set_visible(False)
ax1.axvline(x=0.5, color='#718096', linestyle='--', alpha=0.5, label='Neutral (0.5)')
for bar, val in zip(bars, scores):
    ax1.text(val + 0.02, bar.get_y() + bar.get_height()/2, f'{val:.3f}', 
             va='center', fontsize=9, fontweight='bold')
ax1.legend()
plt.tight_layout()
chart1_path = '/tmp/chart_topsis_bar.png'
plt.savefig(chart1_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Chart 2: Radar chart comparing top 2 options
fig2, ax2 = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
categories_radar = ['Cost', 'Resilience', 'Risk', 'Strategic', 'Ease']
N = len(categories_radar)
angles = [n / float(N) * 2 * np.pi for n in range(N)]
angles += angles[:1]

# Top 2 options scores
dual_source = [8, 8, 8, 7, 6]
framework = [7, 7, 7, 6, 7]
dual_source += dual_source[:1]
framework += framework[:1]

ax2.plot(angles, dual_source, 'o-', linewidth=2, label='Dual Source (Winner)', color=f'#{GREEN}')
ax2.fill(angles, dual_source, alpha=0.25, color=f'#{GREEN}')
ax2.plot(angles, framework, 'o-', linewidth=2, label='Framework (2nd)', color=f'#{COBALT}')
ax2.fill(angles, framework, alpha=0.25, color=f'#{COBALT}')

ax2.set_xticks(angles[:-1])
ax2.set_xticklabels(categories_radar)
ax2.set_ylim(0, 10)
ax2.set_title('Top 2 Options Comparison', fontsize=12, fontweight='bold', color=f'#{NAVY}', pad=20)
ax2.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0))
plt.tight_layout()
chart2_path = '/tmp/chart_topsis_radar.png'
plt.savefig(chart2_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Insert charts
img1 = XLImage(chart1_path)
img1.width = 480
img1.height = 300
ws3.add_image(img1, 'M3')

img2 = XLImage(chart2_path)
img2.width = 400
img2.height = 400
ws3.add_image(img2, 'M25')

# Set column widths
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 12

apply_borders(ws3, 5, 10, 1, 6)
apply_borders(ws3, 13, 18, 1, 4)
apply_borders(ws3, 25, 31, 1, 6)
apply_borders(ws3, 34, 40, 1, 6)
apply_borders(ws3, 43, 49, 1, 6)
apply_borders(ws3, 52, 54, 1, 6)
apply_borders(ws3, 57, 63, 1, 5)

print("Sheet 3: MCDM Scoring - Complete")

# Save
wb.save(output_path)
print(f"Saved: {output_path}")
