import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np

# Define colors
NAVY = "002147"
COBALT = "2B6CB0"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6EAF8"
GREEN = "48BB78"
RED = "E53E3E"
AMBER = "ED8936"

bold_font = Font(bold=True)
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
white_font = Font(bold=True, color=WHITE)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

# Load workbook
output_path = '/Users/jonathonmilne/.openclaw/workspace/category-strategy-process/CATEGORY_STRATEGY_TEMPLATE.xlsx'
wb = load_workbook(output_path)

# ============================================
# SHEET 4: COMPONENT ANALYSIS
# ============================================
ws4 = wb.create_sheet("Component Analysis")

# Title
ws4['A1'] = "COMPONENT ANALYSIS"
ws4['A1'].font = Font(bold=True, size=14, color=NAVY)
ws4.merge_cells('A1:H1')

# Supplier List Section
ws4['A3'] = "SUPPLIER LANDSCAPE"
ws4['A3'].font = Font(bold=True, size=12, color=NAVY)
ws4.merge_cells('A3:H3')

supplier_headers = ['Supplier', 'HQ', 'Revenue (€M)', 'Market Share %', 'Tier', 'Notes']
for col, header in enumerate(supplier_headers, 1):
    cell = ws4.cell(row=4, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Example supplier data
supplier_data = [
    ["Siemens Energy", "Germany", 32000, 18, "Tier 1", "Leading HV equipment manufacturer"],
    ["Hitachi Energy", "Switzerland", 28000, 15, "Tier 1", "Former ABB Power Grids"],
    ["GE Grid Solutions", "USA", 22000, 12, "Tier 1", "Strong in North America"],
    ["Hyosung Heavy Industries", "South Korea", 8500, 8, "Tier 2", "Growing Asian player"],
    ["Toshiba Energy", "Japan", 7200, 6, "Tier 2", "Nuclear & HV focus"],
    ["Mitsubishi Electric", "Japan", 6800, 5, "Tier 2", "Premium positioning"],
]

for row_idx, row_data in enumerate(supplier_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx == 3:
            cell.number_format = '#,##0'
        if col_idx == 4:
            cell.number_format = '0%'

# Should-Cost Components Section
ws4['A13'] = "SHOULD-COST ANALYSIS"
ws4['A13'].font = Font(bold=True, size=12, color=NAVY)
ws4.merge_cells('A13:H13')

shouldcost_headers = ['Component', 'Unit', 'Qty', 'Rate (€)', 'Total (€)', '% of Cost']
for col, header in enumerate(shouldcost_headers, 1):
    cell = ws4.cell(row=14, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# Should-cost breakdown
shouldcost_data = [
    ["Raw Materials", "kg", 15000, 450, 6750000, 0.35],
    ["Labour (Manufacturing)", "hrs", 25000, 85, 2125000, 0.11],
    ["Energy (Electricity)", "MWh", 850, 120, 102000, 0.05],
    ["Overheads", "allocated", 1, 3800000, 3800000, 0.20],
    ["Margin", "%", 0.18, 0, 3420000, 0.18],
    ["Logistics", "shipment", 45, 45000, 2025000, 0.10],
    ["Warranty/Service", "yr", 5, 200000, 1000000, 0.05],
    ["TOTAL", "", "", "", "=SUM(E15:E21)", "=SUM(F15:F21)"],
]

for row_idx, row_data in enumerate(shouldcost_data, 15):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 22:  # Total row
            cell.font = bold_font
            cell.fill = light_blue_fill
        elif row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx in [4, 5] and isinstance(value, (int, float)):
            cell.number_format = '#,##0'
        if col_idx == 6:
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
                cell.number_format = '0%'
            elif isinstance(value, float):
                cell.number_format = '0%'

# Cost Driver Section
ws4['A25'] = "COST DRIVER ANALYSIS (5-YEAR TREND)"
ws4['A25'].font = Font(bold=True, size=12, color=NAVY)
ws4.merge_cells('A25:H25')

driver_headers = ['Driver', '5yr Trend', 'Price Impact %', 'Forward Projection', 'Risk Rating']
for col, header in enumerate(driver_headers, 1):
    cell = ws4.cell(row=26, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center")

# Cost driver data with example indices
driver_data = [
    ["Steel (HRC/CRU Index)", "+45% (2020-2025)", 0.18, "+8% pa next 2yr", "High"],
    ["Copper (LME)", "+38% (2020-2025)", 0.12, "+5% pa next 2yr", "Medium"],
    ["Energy (Nordpool/TTF)", "+62% (2020-2025)", 0.08, "Volatile, +3% pa", "High"],
    ["Labour (Eurostat)", "+18% (2020-2025)", 0.06, "+4% pa next 2yr", "Low"],
    ["Logistics (BDI/Freightos)", "+38% (2020-2025)", 0.05, "Normalizing, -2% pa", "Medium"],
    ["FX (EUR/USD)", "-12% (EUR weaker)", 0.03, "Stable range", "Low"],
]

for row_idx, row_data in enumerate(driver_data, 27):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx == 3 and isinstance(value, float):
            cell.number_format = '0%'
        # Color code risk rating
        if col_idx == 5:
            if value == "High":
                cell.font = Font(color=RED, bold=True)
            elif value == "Medium":
                cell.font = Font(color=AMBER, bold=True)
            elif value == "Low":
                cell.font = Font(color=GREEN, bold=True)

# Create charts for Component Analysis
# Chart 1: Pie chart for should-cost breakdown
fig1, ax1 = plt.subplots(figsize=(7, 7))
components = ['Raw Materials', 'Overheads', 'Margin', 'Labour', 'Logistics', 'Energy', 'Warranty']
values_pie = [35, 20, 18, 11, 10, 5, 5]
colors_pie = [f'#{NAVY}', f'#{COBALT}', f'#{GREEN}', '#718096', '#A0AEC0', '#CBD5E0', '#E2E8F0']
explode = (0.05, 0, 0, 0, 0, 0, 0)  # Explode largest slice
wedges, texts, autotexts = ax1.pie(values_pie, labels=components, autopct='%1.0f%%', 
                                    colors=colors_pie, startangle=90, explode=explode)
for autotext in autotexts:
    autotext.set_color('white')
    autotext.set_fontweight('bold')
    autotext.set_fontsize(9)
ax1.set_title('Should-Cost Component Breakdown', fontsize=12, fontweight='bold', color=f'#{NAVY}')
plt.tight_layout()
chart1_path = '/tmp/chart_shouldcost_pie.png'
plt.savefig(chart1_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Chart 2: Line chart for cost driver trends (2020-2025)
fig2, ax2 = plt.subplots(figsize=(9, 5))
years = [2020, 2021, 2022, 2023, 2024, 2025]
# Normalized indices (2020 = 100)
steel = [100, 145, 178, 165, 142, 145]
energy = [100, 185, 245, 198, 165, 162]
labour = [100, 104, 108, 112, 115, 118]

ax2.plot(years, steel, marker='o', linewidth=2.5, label='Steel (CRU)', color=f'#{NAVY}')
ax2.plot(years, energy, marker='s', linewidth=2.5, label='Energy (Nordpool)', color=f'#{COBALT}')
ax2.plot(years, labour, marker='^', linewidth=2.5, label='Labour (Eurostat)', color=f'#{GREEN}')
ax2.axhline(y=100, color='#718096', linestyle='--', alpha=0.5, label='Baseline (2020)')
ax2.set_xlabel('Year', fontsize=11)
ax2.set_ylabel('Index (2020 = 100)', fontsize=11)
ax2.set_title('Cost Driver Trends (5-Year Lookback)', fontsize=12, fontweight='bold', color=f'#{NAVY}')
ax2.legend(loc='upper left')
ax2.spines['top'].set_visible(False)
ax2.spines['right'].set_visible(False)
ax2.grid(True, alpha=0.3)
ax2.set_ylim(80, 280)

# Add annotations for key changes
ax2.annotate('+45%', xy=(2025, 145), xytext=(2024.5, 155),
            fontsize=9, fontweight='bold', color=f'#{NAVY}')
ax2.annotate('+62%', xy=(2025, 162), xytext=(2024.5, 175),
            fontsize=9, fontweight='bold', color=f'#{COBALT}')

plt.tight_layout()
chart2_path = '/tmp/chart_driver_trends.png'
plt.savefig(chart2_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Insert charts
img1 = XLImage(chart1_path)
img1.width = 420
img1.height = 420
ws4.add_image(img1, 'J3')

img2 = XLImage(chart2_path)
img2.width = 540
img2.height = 300
ws4.add_image(img2, 'J25')

# Set column widths
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 12

apply_borders(ws4, 4, 10, 1, 6)
apply_borders(ws4, 14, 22, 1, 6)
apply_borders(ws4, 26, 32, 1, 5)

print("Sheet 4: Component Analysis - Complete")

# Save
wb.save(output_path)
print(f"Saved: {output_path}")
print("\nAll 5 sheets complete!")
