#!/usr/bin/env python3
"""
CATEGORY_STRATEGY_TEMPLATE.xlsx Generator
Manu Forti Intelligence | March 2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

def create_category_strategy_template():
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    subheader_font = Font(color="FFFFFF", bold=True, size=10)
    instruction_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    instruction_font = Font(color="4A5568", italic=True, size=9)
    input_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    calc_font = Font(color="276749", bold=True)
    
    # ==================== SHEET 0: ASSUMPTIONS ====================
    ws0 = wb.active
    ws0.title = "Assumptions"
    
    ws0['A1'] = "CATEGORY STRATEGY - ASSUMPTIONS & INTAKE"
    ws0['A1'].font = Font(size=16, bold=True, color="002147")
    ws0.merge_cells('A1:F1')
    
    ws0['A2'] = "Complete this sheet with client intake data before proceeding"
    ws0['A2'].font = instruction_font
    ws0['A2'].fill = instruction_fill
    ws0.merge_cells('A2:F2')
    
    ws0['A4'] = "CLIENT INFORMATION"
    ws0['A4'].font = header_font
    ws0['A4'].fill = header_fill
    ws0.merge_cells('A4:F4')
    
    fields = [
        ("Client Name:", "[Enter client name]"),
        ("Category Name:", "[e.g., High Voltage Electrical Equipment]"),
        ("Annual Spend:", "[e.g., 15000000]"),
        ("Currency:", "EUR"),
        ("Timeline Constraint:", "[e.g., DG2 gate in 6 months]"),
        ("Project Pipeline:", "[e.g., 9 projects/yr: 2 large, 3 mid, 4 small]"),
    ]
    
    row = 5
    for label, placeholder in fields:
        ws0[f'A{row}'] = label
        ws0[f'A{row}'].font = Font(bold=True)
        ws0[f'B{row}'] = placeholder
        ws0[f'B{row}'].fill = input_fill
        ws0.merge_cells(f'B{row}:F{row}')
        row += 1
    
    row += 1
    ws0[f'A{row}'] = "INCUMBENT SUPPLIERS"
    ws0[f'A{row}'].font = header_font
    ws0[f'A{row}'].fill = header_fill
    ws0.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws0[f'A{row}'] = "Supplier Name"
    ws0[f'B{row}'] = "Relationship Length"
    ws0[f'C{row}'] = "Spend Share (%)"
    ws0[f'D{row}'] = "Key Issues"
    for col in ['A', 'B', 'C', 'D']:
        ws0[f'{col}{row}'].font = subheader_font
        ws0[f'{col}{row}'].fill = subheader_fill
    row += 1
    
    for i in range(5):
        ws0[f'A{row}'] = f"[Supplier {i+1}]"
        ws0[f'B{row}'] = "[e.g., 3 years]"
        ws0[f'C{row}'] = "[e.g., 40%]"
        ws0[f'D{row}'] = "[Key issues]"
        row += 1
    
    row += 1
    ws0[f'A{row}'] = "KEY PAIN POINTS"
    ws0[f'A{row}'].font = header_font
    ws0[f'A{row}'].fill = header_fill
    ws0.merge_cells(f'A{row}:F{row}')
    row += 1
    
    pain_points = ["Long lead times", "Single source risk", "Cost escalation", "Quality issues", "Other"]
    for i, pp in enumerate(pain_points):
        ws0[f'A{row}'] = f"{i+1}."
        ws0[f'B{row}'] = pp
        ws0[f'C{row}'] = "[Severity: High/Med/Low]"
        ws0[f'D{row}'] = "[Annual cost impact €]"
        row += 1
    
    row += 1
    ws0[f'A{row}'] = "STRATEGIC PRIORITIES"
    ws0[f'A{row}'].font = header_font
    ws0[f'A{row}'].fill = header_fill
    ws0.merge_cells(f'A{row}:F{row}')
    row += 1
    
    priorities = ["Reduce cost", "Improve resilience", "ESG compliance", "Risk mitigation", "Innovation access"]
    for i, p in enumerate(priorities):
        ws0[f'A{row}'] = f"{i+1}."
        ws0[f'B{row}'] = p
        ws0[f'C{row}'] = "[Priority: High/Med/Low]"
        ws0[f'D{row}'] = "[Target outcome]"
        row += 1
    
    row += 2
    ws0[f'A{row}'] = "EVALUATION CRITERIA & WEIGHTS (AHP)"
    ws0[f'A{row}'].font = header_font
    ws0[f'A{row}'].fill = header_fill
    ws0.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws0[f'A{row}'] = "Criterion"
    ws0[f'B{row}'] = "Default Weight"
    ws0[f'C{row}'] = "Client Weight"
    ws0[f'D{row}'] = "Rationale"
    for col in ['A', 'B', 'C', 'D']:
        ws0[f'{col}{row}'].font = subheader_font
        ws0[f'{col}{row}'].fill = subheader_fill
    row += 1
    
    criteria_defaults = [
        ("Cost Reduction", 0.30, "TCO, unit price, lifecycle cost"),
        ("Supply Resilience", 0.25, "Lead time, supply security, dual-source"),
        ("Risk Reduction", 0.20, "Financial, geopolitical, ESG, concentration"),
        ("Strategic Alignment", 0.15, "Partnership depth, innovation, ESG goals"),
        ("Implementation Ease", 0.10, "Speed to implement, change burden"),
    ]
    
    for criterion, default, rationale in criteria_defaults:
        ws0[f'A{row}'] = criterion
        ws0[f'B{row}'] = default
        ws0[f'B{row}'].number_format = '0%'
        ws0[f'C{row}'] = f"=B{row}"
        ws0[f'C{row}'].number_format = '0%'
        ws0[f'C{row}'].fill = input_fill
        ws0[f'D{row}'] = rationale
        row += 1
    
    ws0[f'A{row}'] = "Total Weight:"
    ws0[f'A{row}'].font = Font(bold=True)
    ws0[f'B{row}'] = f"=SUM(B{row-5}:B{row-1})"
    ws0[f'B{row}'].number_format = '0%'
    ws0[f'C{row}'] = f"=SUM(C{row-5}:C{row-1})"
    ws0[f'C{row}'].number_format = '0%'
    ws0[f'C{row}'].font = calc_font
    
    ws0.column_dimensions['A'].width = 25
    ws0.column_dimensions['B'].width = 18
    ws0.column_dimensions['C'].width = 18
    ws0.column_dimensions['D'].width = 40
    
    wb.save('CATEGORY_STRATEGY_TEMPLATE.xlsx')
    print("Created CATEGORY_STRATEGY_TEMPLATE.xlsx with 5 sheets")

if __name__ == "__main__":
    create_category_strategy_template()
