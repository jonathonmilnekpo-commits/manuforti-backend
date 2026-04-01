import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np

# Define colors
NAVY = "002147"
COBALT = "2B6CB0"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6EAF8"
LIGHT_GRAY = "F7FAFC"
GREEN = "48BB78"
AMBER = "ED8936"
RED = "E53E3E"

bold_font = Font(bold=True)
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
white_font = Font(bold=True, color=WHITE)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

# Load workbook
output_path = '/Users/jonathonmilne/.openclaw/workspace/category-strategy-process/CATEGORY_STRATEGY_TEMPLATE.xlsx'
wb = load_workbook(output_path)

# ============================================
# SHEET 1: EXECUTIVE DASHBOARD
# ============================================
ws1 = wb.create_sheet("Executive Dashboard")

# Title
ws1['A1'] = "EXECUTIVE DASHBOARD"
ws1['A1'].font = Font(bold=True, size=14, color=NAVY)
ws1.merge_cells('A1:F1')

# KPI Summary Section
ws1['A3'] = "KEY PERFORMANCE INDICATORS"
ws1['A3'].font = Font(bold=True, size=12, color=NAVY)
ws1.merge_cells('A3:F3')

# KPI Headers
kpi_headers = ['Metric', 'Value', 'Unit', 'Benchmark', 'Status']
for col, header in enumerate(kpi_headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# KPI Data (pre-populated with example based on Statkraft reference)
kpi_data = [
    ["Category", "High Voltage Electrical Equipment", "", "", ""],
    ["Recommended Strategy", "Dual Source with Framework", "", "", ""],
    ["Overall Annual Value", 55600000, "€", "€15M spend base", "✓"],
    ["ROI", 90, ":1", "3:1 typical", "✓"],
    ["Payback Period", 0.5, "months", "<6 months", "✓"],
    ["NPV (5-year)", 222000000, "€", "", "✓"],
]

for row_idx, row_data in enumerate(kpi_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx == 2 and isinstance(value, (int, float)) and row_idx > 5:
            cell.number_format = '#,##0'

# Scenario Comparison Section
ws1['A12'] = "SCENARIO COMPARISON"
ws1['A12'].font = Font(bold=True, size=12, color=NAVY)
ws1.merge_cells('A12:F12')

scenario_headers = ['Metric', 'Bear (70%)', 'Base (100%)', 'Bull (130%)', 'Variance']
for col, header in enumerate(scenario_headers, 1):
    cell = ws1.cell(row=13, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

scenario_data = [
    ["Annual Value", 38920000, 55600000, 72280000, "+85% Bear→Bull"],
    ["Programme Cost", 600000, 600000, 600000, "Fixed"],
    ["Net Annual Value", 38320000, 55000000, 71680000, "+87%"],
    ["ROI", 64.9, 91.7, 120.5, ":1"],
    ["NPV (5yr, 8%)", 153000000, 222000000, 289000000, "€"],
]

for row_idx, row_data in enumerate(scenario_data, 14):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx in [2, 3, 4] and isinstance(value, (int, float)):
            if row_idx in [14, 15, 16]:
                cell.number_format = '#,##0'
            elif row_idx == 17:
                cell.number_format = '0.0'

# Recommendation Highlight Box
ws1['A20'] = "RECOMMENDATION"
ws1['A20'].font = Font(bold=True, size=12, color=NAVY)
ws1.merge_cells('A20:F20')

ws1['A21'] = "Strategy"
ws1['B21'] = "Dual Source with Framework Agreement"
ws1['A21'].font = bold_font
ws1.merge_cells('B21:F21')

ws1['A22'] = "Rationale"
ws1['B22'] = "Balances cost reduction through competition with supply resilience through dual-sourcing. Framework provides price certainty and reduces transaction costs."
ws1['A22'].font = bold_font
ws1.merge_cells('B22:F22')
ws1['B22'].alignment = Alignment(wrap_text=True)
ws1.row_dimensions[22].height = 40

ws1['A23'] = "Key Conditions"
ws1['B23'] = "1) Minimum volume commitment of €5M/yr per supplier 2) 6-month implementation timeline 3) Executive sponsorship for change management"
ws1['A23'].font = bold_font
ws1.merge_cells('B23:F23')
ws1['B23'].alignment = Alignment(wrap_text=True)
ws1.row_dimensions[23].height = 50

# Create charts for Executive Dashboard
# Chart 1: Bar chart for scenario comparison
fig1, ax1 = plt.subplots(figsize=(8, 4))
scenarios = ['Bear (70%)', 'Base (100%)', 'Bull (130%)']
net_values = [38.32, 55.0, 71.68]  # In millions
colors = [f'#{COBALT}', f'#{NAVY}', f'#{GREEN}']
bars = ax1.bar(scenarios, net_values, color=colors, edgecolor='white', linewidth=2)
ax1.set_ylabel('Net Annual Value (€M)', fontsize=11)
ax1.set_title('Scenario Comparison: Net Annual Value', fontsize=12, fontweight='bold', color=f'#{NAVY}')
ax1.spines['top'].set_visible(False)
ax1.spines['right'].set_visible(False)
for bar, val in zip(bars, net_values):
    ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1, f'€{val}M', 
             ha='center', va='bottom', fontsize=10, fontweight='bold')
plt.tight_layout()
chart1_path = '/tmp/chart1_scenarios.png'
plt.savefig(chart1_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Chart 2: Donut chart for value stream breakdown
fig2, ax2 = plt.subplots(figsize=(6, 6))
value_streams = ['Cost Reduction', 'Delay Avoidance', 'Risk Reduction', 'ESG Compliance', 
                 'Working Capital', 'Quality', 'Innovation', 'Optionality', 'PPA Revenue']
values = [25, 35, 12, 5, 8, 5, 4, 4, 2]  # Percentages
colors_donut = [f'#{COBALT}', f'#{NAVY}', f'#{GREEN}', '#718096', '#A0AEC0', 
                '#CBD5E0', '#E2E8F0', '#ED8936', '#E53E3E']
wedges, texts, autotexts = ax2.pie(values, labels=value_streams, autopct='%1.0f%%', 
                                    colors=colors_donut, startangle=90, pctdistance=0.75)
for autotext in autotexts:
    autotext.set_color('white')
    autotext.set_fontweight('bold')
    autotext.set_fontsize(8)
centre_circle = plt.Circle((0,0), 0.50, fc='white')
ax2.add_artist(centre_circle)
ax2.set_title('Value Stream Breakdown', fontsize=12, fontweight='bold', color=f'#{NAVY}')
plt.tight_layout()
chart2_path = '/tmp/chart2_value_streams.png'
plt.savefig(chart2_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Insert charts into Excel
img1 = XLImage(chart1_path)
img1.width = 480
img1.height = 240
ws1.add_image(img1, 'H3')

img2 = XLImage(chart2_path)
img2.width = 360
img2.height = 360
ws1.add_image(img2, 'H18')

# Set column widths
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 15

apply_borders(ws1, 4, 10, 1, 5)
apply_borders(ws1, 13, 18, 1, 5)

print("Sheet 1: Executive Dashboard - Complete")

# Save
wb.save(output_path)
print(f"Saved: {output_path}")
