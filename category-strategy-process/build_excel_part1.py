import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import io
import os

# Create workbook
wb = Workbook()

# Define colors
NAVY = "002147"
COBALT = "2B6CB0"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6EAF8"
LIGHT_GRAY = "F7FAFC"
GREEN = "48BB78"
AMBER = "ED8936"
RED = "E53E3E"

# Define styles
navy_header = NamedStyle(name="navy_header")
navy_header.font = Font(bold=True, color=WHITE, size=11)
navy_header.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
navy_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

light_blue_row = NamedStyle(name="light_blue_row")
light_blue_row.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Helper to apply borders
def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

# ============================================
# SHEET 0: ASSUMPTIONS
# ============================================
ws0 = wb.active
ws0.title = "Assumptions"

# Register styles
wb.add_named_style(navy_header)
wb.add_named_style(light_blue_row)

# Title
ws0['A1'] = "CATEGORY STRATEGY - INPUT ASSUMPTIONS"
ws0['A1'].font = Font(bold=True, size=14, color=NAVY)
ws0.merge_cells('A1:D1')

# Client Information Section
ws0['A3'] = "CLIENT INFORMATION"
ws0['A3'].font = Font(bold=True, size=12, color=NAVY)
ws0.merge_cells('A3:D3')

# Headers
headers = ['Field', 'Value', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws0.cell(row=4, column=col, value=header)
    cell.style = "navy_header"

# Input fields
input_data = [
    ["Client Name", "[Enter Client Name]", "e.g., Statkraft AS"],
    ["Category Name", "[Enter Category]", "e.g., High Voltage Electrical Equipment"],
    ["Annual Spend (€)", "[Enter Amount]", "e.g., 15,000,000"],
    ["Currency", "EUR", "EUR/USD/GBP/NOK"],
    ["Analysis Date", "[Enter Date]", "e.g., 2026-03-15"],
    ["Timeline Constraint", "[Enter Constraint]", "e.g., DG2 gate in 6 months"],
    ["Tier Selection", "Full", "Full / Bundle / Partner"],
]

for row_idx, row_data in enumerate(input_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws0.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.style = "light_blue_row"

# Incumbent Suppliers Section
ws0['A13'] = "INCUMBENT SUPPLIERS"
ws0['A13'].font = Font(bold=True, size=12, color=NAVY)
ws0.merge_cells('A13:D13')

ws0['A14'] = "Supplier Name"
ws0['B14'] = "Relationship Length"
ws0['C14'] = "Annual Spend (€)"
ws0['D14'] = "Notes"
for col in range(1, 5):
    ws0.cell(row=14, column=col).style = "navy_header"

suppliers = [
    ["[Supplier 1]", "[Years]", "[Amount]", "[Notes]"],
    ["[Supplier 2]", "[Years]", "[Amount]", "[Notes]"],
    ["[Supplier 3]", "[Years]", "[Amount]", "[Notes]"],
]
for row_idx, row_data in enumerate(suppliers, 15):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws0.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 1:
            cell.style = "light_blue_row"

# Key Pain Points Section
ws0['A19'] = "KEY PAIN POINTS"
ws0['A19'].font = Font(bold=True, size=12, color=NAVY)
ws0.merge_cells('A19:D19')

ws0['A20'] = "Pain Point"
ws0['B20'] = "Impact (€/yr)"
ws0['C20'] = "Frequency"
ws0['D20'] = "Notes"
for col in range(1, 5):
    ws0.cell(row=20, column=col).style = "navy_header"

pain_points = [
    ["[Pain Point 1]", "[Impact]", "[Freq]", "[Notes]"],
    ["[Pain Point 2]", "[Impact]", "[Freq]", "[Notes]"],
    ["[Pain Point 3]", "[Impact]", "[Freq]", "[Notes]"],
]
for row_idx, row_data in enumerate(pain_points, 21):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws0.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.style = "light_blue_row"

# Strategic Priorities Section
ws0['A25'] = "STRATEGIC PRIORITIES"
ws0['A25'].font = Font(bold=True, size=12, color=NAVY)
ws0.merge_cells('A25:D25')

ws0['A26'] = "Priority"
ws0['B26'] = "Weight %"
ws0['C26'] = "Target Outcome"
ws0['D26'] = "Notes"
for col in range(1, 5):
    ws0.cell(row=26, column=col).style = "navy_header"

priorities = [
    ["[Priority 1]", "[%]", "[Outcome]", "[Notes]"],
    ["[Priority 2]", "[%]", "[Outcome]", "[Notes]"],
    ["[Priority 3]", "[%]", "[Outcome]", "[Notes]"],
]
for row_idx, row_data in enumerate(priorities, 27):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws0.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.style = "light_blue_row"

# Evaluation Criteria Section (AHP Default Weights)
ws0['A31'] = "EVALUATION CRITERIA (AHP WEIGHTS)"
ws0['A31'].font = Font(bold=True, size=12, color=NAVY)
ws0.merge_cells('A31:D31')

ws0['A32'] = "Criterion"
ws0['B32'] = "Weight %"
ws0['C32'] = "Rationale"
ws0['D32'] = "Customizable"
for col in range(1, 5):
    ws0.cell(row=32, column=col).style = "navy_header"

criteria = [
    ["Cost Reduction", 30, "TCO, unit price, lifecycle cost", "Yes"],
    ["Supply Resilience", 25, "Lead time, supply security, dual-source", "Yes"],
    ["Risk Reduction", 20, "Financial, geopolitical, ESG, concentration", "Yes"],
    ["Strategic Alignment", 15, "Partnership depth, innovation, ESG goals", "Yes"],
    ["Implementation Ease", 10, "Speed to implement, change burden", "Yes"],
]
for row_idx, row_data in enumerate(criteria, 33):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws0.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.style = "light_blue_row"
        if col_idx == 2:  # Weight column
            cell.number_format = '0%'

# Set column widths
ws0.column_dimensions['A'].width = 25
ws0.column_dimensions['B'].width = 20
ws0.column_dimensions['C'].width = 35
ws0.column_dimensions['D'].width = 20

# Apply borders
apply_borders(ws0, 4, 11, 1, 4)
apply_borders(ws0, 14, 17, 1, 4)
apply_borders(ws0, 20, 23, 1, 4)
apply_borders(ws0, 26, 29, 1, 4)
apply_borders(ws0, 32, 37, 1, 4)

print("Sheet 0: Assumptions - Complete")

# Save workbook
output_path = '/Users/jonathonmilne/.openclaw/workspace/category-strategy-process/CATEGORY_STRATEGY_TEMPLATE.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
