import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np

# Define colors
NAVY = "002147"
COBALT = "2B6CB0"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6EAF8"
GREEN = "48BB78"
RED = "E53E3E"

bold_font = Font(bold=True)
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
white_font = Font(bold=True, color=WHITE)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

# Load workbook
output_path = '/Users/jonathonmilne/.openclaw/workspace/category-strategy-process/CATEGORY_STRATEGY_TEMPLATE.xlsx'
wb = load_workbook(output_path)

# ============================================
# SHEET 2: FINANCIAL CALCULATOR
# ============================================
ws2 = wb.create_sheet("Financial Calculator")

# Title
ws2['A1'] = "FINANCIAL CALCULATOR"
ws2['A1'].font = Font(bold=True, size=14, color=NAVY)
ws2.merge_cells('A1:H1')

# Value Streams Section
ws2['A3'] = "VALUE STREAMS"
ws2['A3'].font = Font(bold=True, size=12, color=NAVY)
ws2.merge_cells('A3:H3')

value_headers = ['Value Stream', 'Description', 'Annual Value (€)', 'Confidence %', 'Weighted Value (€)']
for col, header in enumerate(value_headers, 1):
    cell = ws2.cell(row=4, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Pre-populated value streams with example data
value_streams_data = [
    ["Cost Reduction", "Price negotiation, volume discounts, TCO optimization", 12500000, 0.75, "=C5*D5"],
    ["Delay Avoidance", "Project schedule protection, liquidated damages avoided", 35000000, 0.65, "=C6*D6"],
    ["PPA Revenue", "Power purchase agreement optimization", 800000, 0.80, "=C7*D7"],
    ["Risk Reduction", "Supply disruption mitigation, insurance cost reduction", 4200000, 0.70, "=C8*D8"],
    ["ESG Compliance", "CSRD alignment, carbon cost avoidance", 1500000, 0.60, "=C9*D9"],
    ["Working Capital", "Payment terms extension, inventory optimization", 800000, 0.75, "=C10*D10"],
    ["Quality/Performance", "Defect reduction, warranty cost avoidance", 500000, 0.65, "=C11*D11"],
    ["Innovation Access", "Early access to new technology, co-development", 200000, 0.50, "=C12*D12"],
    ["Strategic Optionality", "Flexibility value, future sourcing options", 400000, 0.40, "=C13*D13"],
]

for row_idx, row_data in enumerate(value_streams_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx in [3, 5]:
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.number_format = '#,##0'
        if col_idx == 4:
            cell.number_format = '0%'

# Total row
ws2['A14'] = "TOTAL VALUE"
ws2['A14'].font = bold_font
ws2['C14'] = "=SUM(C5:C13)"
ws2['C14'].number_format = '#,##0'
ws2['E14'] = "=SUM(E5:E13)"
ws2['E14'].number_format = '#,##0'
for col in [1, 3, 5]:
    ws2.cell(row=14, column=col).fill = light_blue_fill

# Programme Costs Section
ws2['A16'] = "PROGRAMME COSTS"
ws2['A16'].font = Font(bold=True, size=12, color=NAVY)
ws2.merge_cells('A16:H16')

cost_headers = ['Cost Category', 'Description', 'Year 1', 'Annual (Y2-5)', 'Total 5-Year']
for col, header in enumerate(cost_headers, 1):
    cell = ws2.cell(row=17, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

cost_data = [
    ["Implementation", "Setup, negotiation, contract execution", 300000, 0, 300000],
    ["Annual Running", "Management, monitoring, compliance", 60000, 60000, 300000],
    ["Opportunity Cost", "Internal resource time", 0, 0, 0],
    ["TOTAL", "", "=SUM(C18:C20)", "=SUM(D18:D20)", "=SUM(E18:E20)"],
]

for row_idx, row_data in enumerate(cost_data, 18):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 21:  # Total row
            cell.font = bold_font
            cell.fill = light_blue_fill
        if col_idx in [3, 4, 5] and isinstance(value, str) and value.startswith('='):
            cell.value = value
        elif col_idx in [3, 4, 5]:
            cell.number_format = '#,##0'

# Scenario Analysis Section
ws2['A23'] = "SCENARIO ANALYSIS"
ws2['A23'].font = Font(bold=True, size=12, color=NAVY)
ws2.merge_cells('A23:H23')

scenario_headers2 = ['Metric', 'Bear (70%)', 'Base (100%)', 'Bull (130%)', 'Formula']
for col, header in enumerate(scenario_headers2, 1):
    cell = ws2.cell(row=24, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

scenario_calc_data = [
    ["Annual Gross Value", "=E14*0.7", "=E14", "=E14*1.3", "Base × factor"],
    ["Annual Programme Cost", "=E21", "=E21", "=E21", "Fixed"],
    ["Annual Net Value", "=B25-B26", "=C25-C26", "=D25-D26", "Gross - Cost"],
    ["5-Year NPV (8%)", "=PV(0.08,5,-B27)", "=PV(0.08,5,-C27)", "=PV(0.08,5,-D27)", "PV calculation"],
    ["ROI (Net/Cost)", "=B27/E21", "=C27/E21", "=D27/E21", ":1 ratio"],
    ["Payback (months)", "=(E21/B27)*12", "=(E21/C27)*12", "=(E21/D27)*12", "months"],
]

for row_idx, row_data in enumerate(scenario_calc_data, 25):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx in [2, 3, 4] and isinstance(value, str) and value.startswith('='):
            cell.value = value
            if row_idx in [25, 26, 27]:
                cell.number_format = '#,##0'
            elif row_idx == 28:
                cell.number_format = '#,##0'
            elif row_idx == 29:
                cell.number_format = '0.0'
            elif row_idx == 30:
                cell.number_format = '0.0'

# Summary Section
ws2['A32'] = "FINANCIAL SUMMARY"
ws2['A32'].font = Font(bold=True, size=12, color=NAVY)
ws2.merge_cells('A32:H32')

summary_headers = ['Metric', 'Value', 'Notes']
for col, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=33, column=col, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

summary_data = [
    ["Base Case Annual Net Value", "=C27", "Central estimate"],
    ["5-Year NPV (8% discount)", "=C28", "Present value of benefits"],
    ["ROI Ratio", "=C29", "Return on investment"],
    ["Payback Period", "=C30", "Months to recover costs"],
    ["Break-even Threshold", "=E21/(C27/15)", "Minimum projects to justify"],
]

for row_idx, row_data in enumerate(summary_data, 34):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill
        if col_idx == 2 and isinstance(value, str) and value.startswith('='):
            cell.value = value

# Create waterfall chart for value build-up
fig3, ax3 = plt.subplots(figsize=(10, 5))
categories = ['Baseline\n(0)', 'Cost\nReduction', 'Delay\nAvoidance', 'PPA\nRevenue', 
              'Risk\nReduction', 'ESG', 'Working\nCapital', 'Quality', 'Innovation', 
              'Optionality', 'Programme\nCosts', 'Net\nValue']
# Waterfall values (cumulative)
waterfall_values = [0, 9.4, 22.8, 0.6, 2.9, 0.9, 0.6, 0.3, 0.1, 0.2, -0.6, 37.2]
waterfall_colors = ['#718096', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{GREEN}', f'#{RED}', f'#{NAVY}']

bars = ax3.bar(categories, waterfall_values, color=waterfall_colors, edgecolor='white', linewidth=1)
ax3.axhline(y=0, color='black', linewidth=0.5)
ax3.set_ylabel('Cumulative Value (€M)', fontsize=11)
ax3.set_title('Value Build-up Waterfall', fontsize=12, fontweight='bold', color=f'#{NAVY}')
ax3.spines['top'].set_visible(False)
ax3.spines['right'].set_visible(False)
plt.xticks(rotation=45, ha='right')
plt.tight_layout()
chart3_path = '/tmp/chart3_waterfall.png'
plt.savefig(chart3_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Create 5-year NPV curve
fig4, ax4 = plt.subplots(figsize=(8, 4))
years = [0, 1, 2, 3, 4, 5]
npv_bear = [0, 38.3, 73.8, 106.6, 137.0, 165.0]
npv_base = [0, 55.0, 105.9, 153.0, 196.7, 237.0]
npv_bull = [0, 71.7, 138.0, 199.4, 256.3, 309.0]

ax4.plot(years, npv_bear, marker='o', linewidth=2, label='Bear (70%)', color=f'#{COBALT}')
ax4.plot(years, npv_base, marker='s', linewidth=2, label='Base (100%)', color=f'#{NAVY}')
ax4.plot(years, npv_bull, marker='^', linewidth=2, label='Bull (130%)', color=f'#{GREEN}')
ax4.fill_between(years, npv_bear, npv_bull, alpha=0.2, color=f'#{COBALT}')
ax4.set_xlabel('Year', fontsize=11)
ax4.set_ylabel('Cumulative NPV (€M)', fontsize=11)
ax4.set_title('5-Year NPV Trajectory', fontsize=12, fontweight='bold', color=f'#{NAVY}')
ax4.legend(loc='upper left')
ax4.spines['top'].set_visible(False)
ax4.spines['right'].set_visible(False)
ax4.grid(True, alpha=0.3)
plt.tight_layout()
chart4_path = '/tmp/chart4_npv.png'
plt.savefig(chart4_path, dpi=100, bbox_inches='tight', facecolor='white')
plt.close()

# Insert charts
img3 = XLImage(chart3_path)
img3.width = 560
img3.height = 280
ws2.add_image(img3, 'J3')

img4 = XLImage(chart4_path)
img4.width = 480
img4.height = 240
ws2.add_image(img4, 'J22')

# Set column widths
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 18

apply_borders(ws2, 4, 14, 1, 5)
apply_borders(ws2, 17, 21, 1, 5)
apply_borders(ws2, 24, 30, 1, 5)
apply_borders(ws2, 33, 38, 1, 3)

print("Sheet 2: Financial Calculator - Complete")

# Save
wb.save(output_path)
print(f"Saved: {output_path}")
